"""
School Manager - Flask API Backend (v7 - Full Multi-Tenant)
"""
import hashlib, os, tempfile, secrets, json, re, io, base64, time, hmac, uuid
import requests
import psycopg2, psycopg2.extras
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta


try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    print("WARNING: python-dotenv not installed — .env file will NOT be loaded. "
          "Run: pip install python-dotenv")

app = Flask(__name__)
CORS(app)

from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from functools import wraps
from flask import g

SECRET_KEY = os.environ.get("SECRET_KEY")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY env var is required. Generate one with "
                        "`python -c \"import secrets;print(secrets.token_hex(32))\"` and set it.")
_serializer = URLSafeTimedSerializer(SECRET_KEY, salt="schoolmanager-session")
SESSION_MAX_AGE = 12 * 3600  # 12h — matches a normal school work day

def issue_token(username, school_id, role, student_id=None, is_class_teacher=False, class_id=None, stream_id=None):
    return _serializer.dumps({
        "username": username, "school_id": school_id, "role": role,
        "student_id": student_id, "is_class_teacher": is_class_teacher,
        "class_id": class_id, "stream_id": stream_id,
    })

def require_auth(f):
    """Verifies the Bearer token and sets g.username / g.school_id / g.role etc.
    school_id now comes ONLY from the signed token — never from a header the
    client can freely change."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get("Authorization", "")
        token = auth[7:].strip() if auth.startswith("Bearer ") else request.args.get("token")
        if not token:
            return jsonify({"ok": False, "error": "Authentication required"}), 401
        try:
            data = _serializer.loads(token, max_age=SESSION_MAX_AGE)
        except SignatureExpired:
            return jsonify({"ok": False, "error": "Session expired, please log in again"}), 401
        except BadSignature:
            return jsonify({"ok": False, "error": "Invalid session"}), 401
        g.username = data["username"]; g.school_id = data["school_id"]; g.role = data["role"]
        g.student_id = data.get("student_id"); g.is_class_teacher = data.get("is_class_teacher", False)
        g.class_id = data.get("class_id"); g.stream_id = data.get("stream_id")
        return f(*args, **kwargs)
    return wrapper

def require_role(*roles):
    def deco(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if g.role not in roles:
                return jsonify({"ok": False, "error": "Forbidden"}), 403
            return f(*args, **kwargs)
        return wrapper
    return deco

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL  = os.environ.get("DATABASE_URL", "")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "storage", "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
IMPORT_EXPORT_FOLDER = os.path.join(BASE_DIR, "storage", "import_exports")
os.makedirs(IMPORT_EXPORT_FOLDER, exist_ok=True)
ALLOWED_LOGO_EXT = {"png","jpg","jpeg","gif","webp","svg"}
def _mime_for_ext(ext):
    return {"png":"image/png","jpg":"image/jpeg","jpeg":"image/jpeg",
            "gif":"image/gif","webp":"image/webp","svg":"image/svg+xml"}.get(ext, "application/octet-stream")
_FALLBACK_SUBJECTS = [
    "mathematics","physics","chemistry","biology","geography","history","civics",
    "english","literature","kiswahili","bible knowledge","book keeping","commerce",
    "business studies","historia ya tanzania na maadili",
]
_FALLBACK_ABBR = {
    "historia ya tanzania na maadili":"HTM","mathematics":"MATH","physics":"PHY",
    "chemistry":"CHEM","biology":"BIO","geography":"GEO","history":"HIST",
    "civics":"CIV","english":"ENG","literature":"LIT","kiswahili":"KIS",
    "bible knowledge":"BK","book keeping":"BKP","commerce":"COM","business studies":"BS",
}
_FALLBACK_GRADES = [
    {"min_score":80,"max_score":100,"grade":"A","points":1},{"min_score":70,"max_score":79,"grade":"B","points":2},
    {"min_score":60,"max_score":69,"grade":"C","points":3},{"min_score":50,"max_score":59,"grade":"D","points":4},
    {"min_score":0,"max_score":49,"grade":"F","points":5},
]

# ── SUBSCRIPTIONS (Manual Mobile Money Verification) ───────────
SUBSCRIPTION_PLANS = {
    "free":     {"amount": 0,      "days": 36500,
                 "label": "Free — basic features, forever",
                 "features": ["Marks entry", "On-screen viewing"]},
    "standard": {"amount": 300000, "days": 182,
                 "label": "Standard — 300,000 TZS / 6 months",
                 "features": ["Everything in Free", "Analytics", "PDF downloads", "Announcements", "Results publishing"]},
    "premium":  {"amount": 500000, "days": 365,
                 "label": "Premium — 500,000 TZS / year",
                 "features": ["Everything in Standard", "Priority support"]},
}

def is_subscribed(school_id):
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT subscription_exempt, subscription_status, subscription_expires_at FROM schools WHERE id=%s", (school_id,))
    row = cur.fetchone(); cur.close(); con.close()
    if not row: return False
    exempt, status, expires_at = row
    if exempt: return True
    return bool(status == "active" and expires_at and expires_at > datetime.utcnow())

def _expire_stale_payment_requests(school_id=None):
    """Lazily flips any 'pending' request older than 48h to 'expired'. Called
    at the top of every read/write path that touches payment_requests, so the
    transition happens the moment anyone looks — no scheduler needed. Rows are
    never deleted; only their status changes, preserving the audit trail."""
    con = get_db(); cur = con.cursor()
    if school_id:
        cur.execute("""UPDATE payment_requests SET status='expired'
                       WHERE status='pending' AND school_id=%s
                       AND submitted_at < NOW() - INTERVAL '48 hours'""", (school_id,))
    else:
        cur.execute("""UPDATE payment_requests SET status='expired'
                       WHERE status='pending' AND submitted_at < NOW() - INTERVAL '48 hours'""")
    con.commit(); cur.close(); con.close()

def subscription_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        sid = school_id_from_header()
        if not is_subscribed(sid):
            return jsonify({"ok": False, "error": "subscription_required",
                             "message": "This feature requires an active subscription."}), 402
        return f(*args, **kwargs)
    return wrapper

def _platform_payment_config():
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT business_name, payment_number, networks FROM platform_payment_config WHERE id=1")
    row = cur.fetchone(); cur.close(); con.close()
    if not row: return {"business_name": "", "payment_number": "", "networks": []}
    return {"business_name": row[0], "payment_number": row[1],
            "networks": row[2].split(",") if row[2] else []}




# ── DB ────────────────────────────────────────────────────────
def get_db():
    con = psycopg2.connect(DATABASE_URL)
    con.autocommit = False
    return con

def to_dict(row, cur):
    if row is None: return None
    return dict(zip([d[0] for d in cur.description], row))

def to_dicts(rows, cur):
    if not rows: return []
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in rows]

# ── PASSWORD ──────────────────────────────────────────────────
def hash_password(pw, iterations=260_000):
    salt = secrets.token_hex(16)
    dk   = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), iterations)
    return f"{iterations}:{salt}:{dk.hex()}"

def verify_password(pw, stored):
    try:
        parts = stored.split(":")
        if len(parts) == 3:
            iterations, salt, dk_hex = parts
            iterations = int(iterations)
        else:
            # Legacy hashes (created before iteration count was stored) always used 260,000.
            salt, dk_hex = parts
            iterations = 260_000
        dk = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), iterations)
        return dk.hex() == dk_hex
    except: return False

# ── SCHOOL_ID HELPERS ─────────────────────────────────────────
def school_id_from_header():
    sid = request.headers.get("X-School-ID")
    if not sid:
        # Browser navigation (<a href>, window.open) used for PDF downloads
        # can't set custom headers, so those links pass school_id as a
        # query parameter instead — fall back to that here.
        sid = request.args.get("school_id")
    if sid:
        try: return int(sid)
        except: pass
    return 1

_REG_CODE_RE = re.compile(r'^[A-Za-z0-9_-]{3,32}$')

def valid_reg_code(code):
    return bool(code) and bool(_REG_CODE_RE.match(code.strip()))

def format_student_display_id(school_id, school_student_no):
    if not school_student_no: return str(school_id)
    return f"{int(school_id):05d}/{int(school_student_no):04d}"

def get_school_id_by_reg_code(reg_code):
    """Resolve a school's registration code (case-insensitive) to its school_id.
    This is the single source of truth for which school a login belongs to —
    it replaces guessing based on username/password alone, which is what let
    a teacher with the same username+password in two different schools get
    logged into the wrong one."""
    if not reg_code: return None
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT id FROM schools WHERE LOWER(reg_code)=LOWER(%s)", (reg_code.strip(),))
    row = cur.fetchone(); cur.close(); con.close()
    return row[0] if row else None

def generate_unique_reg_code(base):
    """Auto-generate a unique reg_code from a school name, e.g. for schools
    that don't pick their own, or for backfilling pre-existing schools."""
    base = re.sub(r'[^A-Za-z0-9_-]', '', (base or "").strip().replace(" ", "_"))[:24] or "school"
    con = get_db(); cur = con.cursor()
    candidate = base; suffix = 0
    while True:
        cur.execute("SELECT 1 FROM schools WHERE LOWER(reg_code)=LOWER(%s)", (candidate,))
        if not cur.fetchone(): break
        suffix += 1
        candidate = f"{base}_{suffix}"
    cur.close(); con.close()
    return candidate

def get_config_val(school_id, key, default=""):
    try:
        con = get_db(); cur = con.cursor()
        cur.execute("SELECT value FROM school_config WHERE school_id=%s AND key=%s", (school_id, key))
        row = cur.fetchone(); cur.close(); con.close()
        return row[0] if row else default
    except: return default

def set_config_val(school_id, key, value):
    con = get_db(); cur = con.cursor()
    cur.execute("""INSERT INTO school_config(school_id,key,value) VALUES(%s,%s,%s)
                   ON CONFLICT(school_id,key) DO UPDATE SET value=EXCLUDED.value""",
                (school_id, key, value))
    con.commit(); cur.close(); con.close()

def get_school_name(school_id):
    return get_config_val(school_id, "school_name", "School Name")

def is_registration_complete(school_id):
    return get_config_val(school_id, "registration_complete", "0") == "1"

def get_subjects(school_id):
    try:
        con = get_db(); cur = con.cursor()
        cur.execute("SELECT name FROM school_subjects WHERE school_id=%s ORDER BY sort_order,name", (school_id,))
        rows = cur.fetchall(); cur.close(); con.close()
        if rows: return [r[0] for r in rows]
    except: pass
    return list(_FALLBACK_SUBJECTS)

def get_subject_map(school_id):
    try:
        con = get_db(); cur = con.cursor()
        cur.execute("SELECT name,abbreviation FROM school_subjects WHERE school_id=%s ORDER BY sort_order,name", (school_id,))
        rows = cur.fetchall(); cur.close(); con.close()
        if rows: return {r[0]:r[1] for r in rows}
    except: pass
    return dict(_FALLBACK_ABBR)

def get_grade_rules(school_id):
    try:
        con = get_db(); cur = con.cursor()
        cur.execute("SELECT min_score,max_score,grade,points FROM grade_config WHERE school_id=%s ORDER BY min_score DESC", (school_id,))
        rows = cur.fetchall(); cur.close(); con.close()
        if rows:
            # Grades saved before "points" existed (or left blank by the admin) get a
            # sensible default assigned automatically: best grade = 1 point, next = 2,
            # etc. (same convention NECTA uses) — so division/points never silently
            # fail just because a school never typed points in.
            return [{"min_score":r[0],"max_score":r[1],"grade":r[2],
                     "points": r[3] if r[3] is not None else (i+1)} for i,r in enumerate(rows)]
    except: pass
    return list(_FALLBACK_GRADES)

def get_grade(school_id, score):
    if score is None: return "-"
    for r in get_grade_rules(school_id):
        if score >= r["min_score"]: return r["grade"]
    return "F"

def get_active_term(school_id):
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT * FROM terms WHERE school_id=%s AND status='open' ORDER BY id DESC LIMIT 1", (school_id,))
    row = cur.fetchone(); r = to_dict(row, cur) if row else None
    cur.close(); con.close(); return r

def get_term_by_id(school_id, tid):
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT * FROM terms WHERE id=%s AND school_id=%s", (tid, school_id))
    row = cur.fetchone(); r = to_dict(row, cur) if row else None
    cur.close(); con.close(); return r

def get_students_in_scope(school_id, class_id, stream_id=None):
    con = get_db(); cur = con.cursor()
    if stream_id:
        cur.execute("""SELECT s.id,s.name,s.class_id,s.stream_id,c.class_name,st.stream_name
                       FROM students s JOIN classes c ON s.class_id=c.id
                       LEFT JOIN streams st ON s.stream_id=st.id
                       WHERE s.school_id=%s AND s.class_id=%s AND s.stream_id=%s ORDER BY s.name""",
                    (school_id, class_id, stream_id))
    else:
        cur.execute("""SELECT s.id,s.name,s.class_id,s.stream_id,c.class_name,st.stream_name
                       FROM students s JOIN classes c ON s.class_id=c.id
                       LEFT JOIN streams st ON s.stream_id=st.id
                       WHERE s.school_id=%s AND s.class_id=%s ORDER BY s.name""",
                    (school_id, class_id))
    rows = to_dicts(cur.fetchall(), cur); cur.close(); con.close()
    return rows

def calc_ca_avg(school_id, student_id, subject, term_id):
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT score FROM ca_scores WHERE school_id=%s AND student_id=%s AND subject=%s AND term_id=%s",
                (school_id, student_id, subject, term_id))
    rows = cur.fetchall(); cur.close(); con.close()
    if not rows: return None
    return sum(r[0] for r in rows) / len(rows)

def calc_final(school_id, student_id, subject, term_id):
    term = get_term_by_id(school_id, term_id)
    if not term: return None
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT score FROM exam_scores WHERE school_id=%s AND student_id=%s AND subject=%s AND term_id=%s",
                (school_id, student_id, subject, term_id))
    row = cur.fetchone(); cur.close(); con.close()
    if not row: return None
    ca_avg = calc_ca_avg(school_id, student_id, subject, term_id)
    if ca_avg is None: return None
    return (ca_avg/100)*term["ca_weight"] + (row[0]/100)*term["exam_weight"]

def calc_student_average(school_id, student_id, term_id):
    subjects = get_subjects(school_id)
    finals = [calc_final(school_id, student_id, s, term_id) for s in subjects]
    finals = [f for f in finals if f is not None]
    return sum(finals)/len(finals) if finals else 0

def _assign_positions(rows, key):
    rows.sort(key=lambda x: (x[key] is None, -(x[key] or 0)))
    for i, r in enumerate(rows):
        if i == 0: r["position"] = 1
        elif r[key] == rows[i-1][key]: r["position"] = rows[i-1]["position"]
        else: r["position"] = i+1

def get_ranking(school_id, students, term_id):
    rows = []
    for s in students:
        avg = calc_student_average(school_id, s["id"], term_id)
        rows.append({"id":s["id"],"name":s["name"],"average":round(avg,2),"grade":get_grade(school_id,avg)})
    _assign_positions(rows, "average")
    return rows

def get_positions(school_id, student_id, class_id, stream_id, term_id):
    all_class = get_students_in_scope(school_id, class_id)
    class_ranking = get_ranking(school_id, all_class, term_id)
    c_pos   = next((r["position"] for r in class_ranking if r["id"]==student_id), "-")
    c_total = len(class_ranking)
    s_pos, s_total = None, None
    if stream_id:
        stream_studs   = get_students_in_scope(school_id, class_id, stream_id)
        stream_ranking = get_ranking(school_id, stream_studs, term_id)
        s_pos   = next((r["position"] for r in stream_ranking if r["id"]==student_id), "-")
        s_total = len(stream_ranking)
    return c_pos, c_total, s_pos, s_total

def get_subject_position(school_id, student_id, subject, class_id, stream_id, term_id):
    scope = get_students_in_scope(school_id, class_id, stream_id)
    scores = []
    for s in scope:
        f = calc_final(school_id, s["id"], subject, term_id)
        if f is not None: scores.append({"id":s["id"],"score":f})
    _assign_positions(scores, "score")
    for s in scores:
        if s["id"] == student_id: return s["position"]
    return "-"

def teacher_can_access(school_id, username, subject, class_id, stream_id=None):
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT id FROM subject_assignments
                   WHERE school_id=%s AND username=%s AND subject=%s AND class_id=%s
                   AND (stream_id=%s OR stream_id IS NULL)""",
                (school_id, username, subject, class_id, stream_id))
    row = cur.fetchone(); cur.close(); con.close()
    return row is not None


# ── BULK SCORE FETCHING (performance) ───────────────────────────
# Report cards, their PDFs, and the parent portal used to compute every
# figure (CA average, final score, class position, stream position,
# subject position) with its own tiny query on its own fresh DB
# connection — repeated once per subject, and for positions, once per
# subject PER STUDENT in the class. For a class of 40 students and 15
# subjects that's thousands of round trips for a single report. These
# helpers fetch every CA/exam score for a whole class+term in exactly
# two queries; everything else (averages, finals, rankings) is then
# plain Python math with zero further DB access.


def _final_from_entry(entry, ca_weight, exam_weight):
    if not entry: return None
    ca_vals = list(entry["ca"].values())
    if not ca_vals or entry["exam"] is None: return None
    ca_avg = sum(ca_vals) / len(ca_vals)
    return (ca_avg/100)*ca_weight + (entry["exam"]/100)*exam_weight

def compute_student_finals(scores_bulk, student_id, subjects, ca_weight, exam_weight):
    student_data = scores_bulk.get(student_id, {})
    return {subj: _final_from_entry(student_data.get(subj), ca_weight, exam_weight) for subj in subjects}

def compute_average_from_finals(finals):
    vals = [v for v in finals.values() if v is not None]
    return sum(vals)/len(vals) if vals else 0

def get_subject_rank_map(class_rows, subject):
    """class_rows: [{"id":..., "finals":{subject:final_or_None}}, ...] -> {student_id: position}"""
    scored = [{"id": r["id"], "score": r["finals"].get(subject)} for r in class_rows if r["finals"].get(subject) is not None]
    _assign_positions(scored, "score")
    return {r["id"]: r["position"] for r in scored}

def get_subject_assess_rank_map(scores_bulk, student_ids, subject, assess):
    """Rank by a single assessment's raw score (a CA name, 'exam', or 'test:ID')
    rather than the weighted final."""
    scored = []
    for stid in student_ids:
        entry = scores_bulk.get(stid, {}).get(subject)
        if not entry: continue
        val = _score_for_assess(entry, assess)
        if val is not None: scored.append({"id": stid, "score": val})
    _assign_positions(scored, "score")
    return {r["id"]: r["position"] for r in scored}

def get_class_report_data(school_id, term_id, class_id, stream_id, subjects, ca_weight, exam_weight):
    """
    One-shot computation of everything needed to render a report card /
    PDF / parent result for an entire class (and, if given, a stream
    within it) for one term — exactly 2 score queries total, regardless
    of class size or subject count.
    Returns (class_rows, class_rank_map, stream_rank_map, scores_bulk).
    class_rows: [{"id","name","stream_id","average","finals":{subject:final}}]
    """
    class_students = get_students_in_scope(school_id, class_id)
    ids = [s["id"] for s in class_students]
    scores_bulk = get_term_scores_bulk(school_id, term_id, ids)
    class_rows = []
    for s in class_students:
        finals = compute_student_finals(scores_bulk, s["id"], subjects, ca_weight, exam_weight)
        avg = compute_average_from_finals(finals)
        class_rows.append({"id": s["id"], "name": s["name"], "stream_id": s.get("stream_id"),
                           "average": round(avg, 2), "finals": finals})
    _assign_positions(class_rows, "average")
    class_rank_map = {r["id"]: r for r in class_rows}
    stream_rank_map = None
    if stream_id:
        stream_rows = [dict(r) for r in class_rows if r["stream_id"] == stream_id]
        _assign_positions(stream_rows, "average")
        stream_rank_map = {r["id"]: r for r in stream_rows}
    return class_rows, class_rank_map, stream_rank_map, scores_bulk


def _get_all_terms_ordered(school_id):
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT * FROM terms WHERE school_id=%s ORDER BY id ASC", (school_id,))
    rows = to_dicts(cur.fetchall(), cur); cur.close(); con.close()
    return rows

def _assessments_for_term(school_id, term, class_id=None):
    assessments = [f"CA{i}" for i in range(1, term["ca_count"] + 1)] + ["exam"]
    for t in get_term_tests(school_id, term["id"], class_id):
        assessments.append(f"test:{t['id']}")
    return assessments

def get_all_students_in_school(school_id):
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT s.id,s.name,s.class_id,s.stream_id,c.class_name,st.stream_name
                   FROM students s JOIN classes c ON s.class_id=c.id
                   LEFT JOIN streams st ON s.stream_id=st.id
                   WHERE s.school_id=%s ORDER BY s.name""", (school_id,))
    rows = to_dicts(cur.fetchall(), cur); cur.close(); con.close()
    return rows

def get_term_tests(school_id, term_id, class_id=None):
    """Returns tests for a term. If class_id is given, only tests that apply
    to that class (all_classes=1, or explicitly listed in test_classes) are
    returned — this is what scopes a test to teachers/students in the
    participating classes only. If class_id is None, every test in the term
    is returned (used for admin management and term-wide publishing)."""
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT id,label,all_classes FROM term_tests WHERE school_id=%s AND term_id=%s ORDER BY id",(school_id,term_id))
    rows=cur.fetchall()
    result=[]
    for tid,label,all_classes in rows:
        if class_id is not None and not all_classes:
            cur.execute("SELECT 1 FROM test_classes WHERE test_id=%s AND class_id=%s",(tid,class_id))
            if not cur.fetchone(): continue
        result.append({"id":tid,"label":label,"all_classes":bool(all_classes)})
    cur.close(); con.close()
    return result

def get_test_class_ids(test_id):
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT class_id FROM test_classes WHERE test_id=%s",(test_id,))
    rows=[r[0] for r in cur.fetchall()]
    cur.close(); con.close()
    return rows

def get_term_scores_bulk(school_id, term_id, student_ids):
    if not student_ids: return {}
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT student_id, subject, ca_name, score FROM ca_scores
                   WHERE school_id=%s AND term_id=%s AND student_id=ANY(%s)""",
                (school_id, term_id, student_ids))
    ca_rows = cur.fetchall()
    cur.execute("""SELECT student_id, subject, score FROM exam_scores
                   WHERE school_id=%s AND term_id=%s AND student_id=ANY(%s)""",
                (school_id, term_id, student_ids))
    exam_rows = cur.fetchall()
    cur.execute("""SELECT student_id, subject, test_id, score FROM test_scores
                   WHERE school_id=%s AND term_id=%s AND student_id=ANY(%s)""",
                (school_id, term_id, student_ids))
    test_rows = cur.fetchall()
    cur.close(); con.close()
    data = {}
    for student_id, subject, ca_name, score in ca_rows:
        d = data.setdefault(student_id, {}).setdefault(subject, {"ca": {}, "exam": None, "tests": {}})
        d["ca"][ca_name] = score
    for student_id, subject, score in exam_rows:
        d = data.setdefault(student_id, {}).setdefault(subject, {"ca": {}, "exam": None, "tests": {}})
        d["exam"] = score
    for student_id, subject, test_id, score in test_rows:
        d = data.setdefault(student_id, {}).setdefault(subject, {"ca": {}, "exam": None, "tests": {}})
        d["tests"][test_id] = score
    return data

def _score_for_assess(entry, assess):
    if not entry: return None
    if assess == "exam": return entry.get("exam")
    if isinstance(assess,str) and assess.startswith("test:"):
        try: tid = int(assess.split(":",1)[1])
        except ValueError: return None
        return (entry.get("tests") or {}).get(tid)
    return (entry.get("ca") or {}).get(assess)

def _compute_overall_series(school_id, class_id, stream_id, subjects):
    students = get_students_in_scope(school_id, class_id, stream_id) if class_id \
        else get_all_students_in_school(school_id)
    student_ids = [s["id"] for s in students]
    name_map = {s["id"]: s["name"] for s in students}
    if not student_ids: return [], name_map
    points = []
    for term in _get_all_terms_ordered(school_id):
        scores_bulk = get_term_scores_bulk(school_id, term["id"], student_ids)
        test_label_map = {t["id"]: t["label"] for t in get_term_tests(school_id, term["id"], class_id)}
        for assess in _assessments_for_term(school_id, term, class_id):
            values = {}
            for sid in student_ids:
                entry_map = scores_bulk.get(sid, {})
                vals = []
                for subj in subjects:
                    v = _score_for_assess(entry_map.get(subj), assess)
                    if v is not None: vals.append(v)
                values[sid] = round(sum(vals) / len(vals), 2) if vals else None
            present = {sid: v for sid, v in values.items() if v is not None}
            if not present: continue
            ranked = [{"id": sid, "score": v} for sid, v in present.items()]
            _assign_positions(ranked, "score")
            ranks = {r["id"]: r["position"] for r in ranked}
            avg = round(sum(present.values()) / len(present), 2)
            if assess == "exam": label = f"{term['label']} Exam"
            elif assess.startswith("test:"):
                label = f"{term['label']} {test_label_map.get(int(assess.split(':')[1]),'Test')}"
            else: label = f"{term['label']} {assess}"
            points.append({"term_id": term["id"], "assess": assess, "label": label,
                           "values": values, "avg": avg, "ranks": ranks, "student_count": len(present)})
    return points, name_map

def _compute_subject_series(school_id, class_id, stream_id, subject):
    students = get_students_in_scope(school_id, class_id, stream_id)
    student_ids = [s["id"] for s in students]
    name_map = {s["id"]: s["name"] for s in students}
    if not student_ids: return [], name_map
    points = []
    for term in _get_all_terms_ordered(school_id):
        scores_bulk = get_term_scores_bulk(school_id, term["id"], student_ids)
        test_label_map = {t["id"]: t["label"] for t in get_term_tests(school_id, term["id"], class_id)}
        for assess in _assessments_for_term(school_id, term, class_id):
            values = {}
            for sid in student_ids:
                v = _score_for_assess(scores_bulk.get(sid, {}).get(subject), assess)
                if v is not None: values[sid] = v
            if not values: continue
            ranked = [{"id": sid, "score": v} for sid, v in values.items()]
            _assign_positions(ranked, "score")
            ranks = {r["id"]: r["position"] for r in ranked}
            avg = round(sum(values.values()) / len(values), 2)
            if assess == "exam": label = f"{term['label']} Exam"
            elif assess.startswith("test:"):
                label = f"{term['label']} {test_label_map.get(int(assess.split(':')[1]),'Test')}"
            else: label = f"{term['label']} {assess}"
            points.append({"term_id": term["id"], "assess": assess, "label": label,
                           "values": values, "avg": avg, "ranks": ranks, "student_count": len(values)})
    return points, name_map

@app.route("/api/tests", methods=["GET"])
@require_auth
def api_get_tests():
    sid = school_id_from_header(); term_id = request.args.get("term_id")
    class_id = request.args.get("class_id")
    class_id = int(class_id) if class_id else None
    if not term_id:
        term = get_active_term(sid)
        if not term: return jsonify([])
        term_id = term["id"]
    else: term_id = int(term_id)
    tests = get_term_tests(sid, term_id, class_id)
    if class_id is None:
        # Admin management view — include which classes each test applies to.
        for t in tests:
            t["class_ids"] = [] if t["all_classes"] else get_test_class_ids(t["id"])
    return jsonify(tests)

@app.route("/api/tests", methods=["POST"])
@require_auth
@require_role("admin")
def api_create_test():
    sid = school_id_from_header(); d = request.json or {}
    term_id = d.get("term_id"); label = (d.get("label") or "").strip()
    class_ids = d.get("class_ids") or []
    if not term_id or not label: return jsonify({"ok":False,"error":"term_id and label required"}),400
    all_classes = 0 if class_ids else 1
    con=get_db(); cur=con.cursor()
    cur.execute("INSERT INTO term_tests(school_id,term_id,label,all_classes) VALUES(%s,%s,%s,%s) RETURNING id",
                (sid,int(term_id),label,all_classes))
    new_id=cur.fetchone()[0]
    for cid in class_ids:
        try: cur.execute("INSERT INTO test_classes(test_id,class_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",(new_id,int(cid)))
        except (TypeError,ValueError): pass
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True,"id":new_id})

@app.route("/api/tests/<int:tid>", methods=["DELETE"])
@require_auth
def api_delete_test(tid):
    sid = g.school_id
    con=get_db(); cur=con.cursor()
    cur.execute("DELETE FROM test_scores WHERE school_id=%s AND test_id=%s",(sid,tid))
    cur.execute("DELETE FROM published_assessments WHERE school_id=%s AND assess_key=%s",(sid,f"test:{tid}"))
    cur.execute("DELETE FROM test_classes WHERE test_id=%s",(tid,))
    cur.execute("DELETE FROM term_tests WHERE id=%s AND school_id=%s",(tid,sid))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True})

@app.route("/api/marks/test", methods=["POST"])
@require_auth
def api_enter_test():
    sid = g.school_id; d = request.json
    username = g.username
    subject=d.get("subject","").lower().strip()
    class_id=int(d.get("class_id")); stream_id=d.get("stream_id") or None
    student_id=int(d.get("student_id")); test_id=int(d.get("test_id")); score=float(d.get("score"))
    if not (0<=score<=100): return jsonify({"ok":False,"error":"Score must be 0-100"}),400
    if g.role=="teacher" and not teacher_can_access(sid,username,subject,class_id,stream_id):
        return jsonify({"ok":False,"error":"Access denied"}),403
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT term_id, all_classes FROM term_tests WHERE id=%s AND school_id=%s",(test_id,sid))
    row=cur.fetchone()
    if not row: cur.close(); con.close(); return jsonify({"ok":False,"error":"Test not found"}),404
    term_id, all_classes = row
    if not all_classes:
        cur.execute("SELECT 1 FROM test_classes WHERE test_id=%s AND class_id=%s",(test_id,class_id))
        if not cur.fetchone():
            cur.close(); con.close()
            return jsonify({"ok":False,"error":"This class is not part of this test"}),403
    cur.execute("""INSERT INTO test_scores(school_id,student_id,subject,test_id,score,entered_by,term_id)
                   VALUES(%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT(school_id,student_id,subject,test_id,term_id)
                   DO UPDATE SET score=EXCLUDED.score,entered_by=EXCLUDED.entered_by""",
                (sid,student_id,subject,test_id,score,username,term_id))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

def _build_common_cards(points, name_map, school_id):
    """Shared card logic for both overall (admin/class-teacher) and
    subject-level (subject teacher) analytics. A student is only flagged as
    improved/declining/outstanding/at-risk when BOTH the mark and the
    position move together — this is the whole point of the spec: marks
    alone or position alone can mislead when exam difficulty varies."""
    if not points:
        return {"graph": [], "current_label": None, "average": None,
                "outstanding": [], "needs_support": [], "improved": [],
                "declining": [], "at_risk": []}

    graph = [{"label": p["label"], "value": p["avg"]} for p in points[-5:]]
    latest = points[-1]
    prev = points[-2] if len(points) >= 2 else None
    latest_vals, latest_ranks, latest_avg = latest["values"], latest["ranks"], latest["avg"]
    total = len(latest_ranks)

    outstanding, needs_support = [], []
    for sid, val in latest_vals.items():
        if val is None: continue
        pos = latest_ranks.get(sid)
        if pos is None: continue
        name = name_map.get(sid, "?")
        if pos <= 3 and val >= latest_avg + 5:
            outstanding.append({"id": sid, "name": name, "value": val, "position": pos})
        if total >= 3 and pos > total - 3 and val <= latest_avg - 5:
            needs_support.append({"id": sid, "name": name, "value": val, "position": pos})
    outstanding.sort(key=lambda r: r["position"])
    needs_support.sort(key=lambda r: -r["position"])

    improved, declining = [], []
    if prev:
        prev_vals, prev_ranks = prev["values"], prev["ranks"]
        for sid in set(latest_vals) & set(prev_vals):
            lv, pv = latest_vals[sid], prev_vals[sid]
            lp, pp = latest_ranks.get(sid), prev_ranks.get(sid)
            if None in (lv, pv, lp, pp): continue
            name = name_map.get(sid, "?")
            if lv > pv and lp < pp:
                improved.append({"id": sid, "name": name, "prev_value": pv, "current_value": lv,
                                 "prev_position": pp, "current_position": lp})
            elif lv < pv and lp > pp:
                declining.append({"id": sid, "name": name, "prev_value": pv, "current_value": lv,
                                  "prev_position": pp, "current_position": lp})

    at_risk = []
    grade_rules = get_grade_rules(school_id)
    lowest_rule = min(grade_rules, key=lambda r: r["min_score"]) if grade_rules else None
    if prev and lowest_rule:
        prev_vals = prev["values"]
        for sid in set(latest_vals) & set(prev_vals):
            lv, pv = latest_vals[sid], prev_vals[sid]
            if lv is None or pv is None: continue
            if lowest_rule["min_score"] <= lv <= lowest_rule["max_score"] and \
               lowest_rule["min_score"] <= pv <= lowest_rule["max_score"]:
                at_risk.append({"id": sid, "name": name_map.get(sid, "?"), "value": lv,
                                "reason": "Lowest grade in two consecutive examinations."})

    return {"graph": graph, "current_label": latest["label"], "average": round(latest_avg, 2),
            "outstanding": outstanding, "needs_support": needs_support,
            "improved": improved, "declining": declining, "at_risk": at_risk}

def _best_weakest_subject(school_id, class_id, stream_id, subjects, term_id, assess):
    students = get_students_in_scope(school_id, class_id, stream_id) if class_id \
        else get_all_students_in_school(school_id)
    ids = [s["id"] for s in students]
    if not ids: return None, None
    scores_bulk = get_term_scores_bulk(school_id, term_id, ids)
    subj_avgs = {}
    for subj in subjects:
        vals = []
        for sid in ids:
            entry = scores_bulk.get(sid, {}).get(subj)
            if not entry: continue
            v = entry["exam"] if assess == "exam" else entry["ca"].get(assess)
            if v is not None: vals.append(v)
        if vals: subj_avgs[subj] = round(sum(vals) / len(vals), 2)
    if not subj_avgs: return None, None
    best = max(subj_avgs.items(), key=lambda kv: kv[1])
    weak = min(subj_avgs.items(), key=lambda kv: kv[1])
    return {"subject": best[0], "average": best[1]}, {"subject": weak[0], "average": weak[1]}





# ── INIT DB ───────────────────────────────────────────────────
def init_db():
    con = get_db(); cur = con.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS schools (
        id            SERIAL PRIMARY KEY,
        school_name   TEXT NOT NULL,
        registered_at TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS users (
        id                   SERIAL,
        username             TEXT NOT NULL,
        password             TEXT NOT NULL,
        role                 TEXT NOT NULL DEFAULT 'teacher',
        school_id            INTEGER NOT NULL DEFAULT 1,
        is_class_teacher     INTEGER DEFAULT 0,
        class_id             INTEGER DEFAULT NULL,
        stream_id            INTEGER DEFAULT NULL,
        must_change_password INTEGER DEFAULT 0,
        student_id           INTEGER DEFAULT NULL,
        PRIMARY KEY(username, school_id)
    );
    CREATE TABLE IF NOT EXISTS classes (
        id         SERIAL PRIMARY KEY,
        school_id  INTEGER NOT NULL DEFAULT 1,
        class_name TEXT NOT NULL,
        UNIQUE(school_id, class_name)
    );
    CREATE TABLE IF NOT EXISTS streams (
        id          SERIAL PRIMARY KEY,
        school_id   INTEGER NOT NULL DEFAULT 1,
        class_id    INTEGER NOT NULL,
        stream_name TEXT NOT NULL,
        UNIQUE(class_id, stream_name)
    );
    CREATE TABLE IF NOT EXISTS students (
        id           SERIAL PRIMARY KEY,
        school_id    INTEGER NOT NULL DEFAULT 1,
        name         TEXT NOT NULL,
        class_id     INTEGER NOT NULL,
        stream_id    INTEGER DEFAULT NULL,
        phone_number TEXT DEFAULT NULL
    );
    CREATE TABLE IF NOT EXISTS subject_assignments (
        id        SERIAL PRIMARY KEY,
        school_id INTEGER NOT NULL DEFAULT 1,
        username  TEXT NOT NULL,
        subject   TEXT NOT NULL,
        class_id  INTEGER NOT NULL,
        stream_id INTEGER DEFAULT NULL,
        UNIQUE(school_id, username, subject, class_id, stream_id)
    );
    CREATE TABLE IF NOT EXISTS terms (
        id          SERIAL PRIMARY KEY,
        school_id   INTEGER NOT NULL DEFAULT 1,
        label       TEXT NOT NULL,
        ca_count    INTEGER NOT NULL DEFAULT 2,
        ca_weight   INTEGER NOT NULL DEFAULT 30,
        exam_weight INTEGER NOT NULL DEFAULT 70,
        status      TEXT NOT NULL DEFAULT 'open'
    );
    CREATE TABLE IF NOT EXISTS ca_scores (
        id         SERIAL PRIMARY KEY,
        school_id  INTEGER NOT NULL DEFAULT 1,
        student_id INTEGER NOT NULL,
        subject    TEXT NOT NULL,
        ca_name    TEXT NOT NULL,
        score      REAL NOT NULL,
        entered_by TEXT,
        term_id    INTEGER NOT NULL,
        UNIQUE(school_id, student_id, subject, ca_name, term_id)
    );
    CREATE TABLE IF NOT EXISTS exam_scores (
        id         SERIAL PRIMARY KEY,
        school_id  INTEGER NOT NULL DEFAULT 1,
        student_id INTEGER NOT NULL,
        subject    TEXT NOT NULL,
        score      REAL NOT NULL,
        entered_by TEXT,
        term_id    INTEGER NOT NULL,
        UNIQUE(school_id, student_id, subject, term_id)
    );
    CREATE TABLE IF NOT EXISTS remarks (
        school_id            INTEGER NOT NULL DEFAULT 1,
        student_id           INTEGER NOT NULL,
        term_id              INTEGER NOT NULL,
        class_teacher_remark TEXT DEFAULT '',
        head_remark          TEXT DEFAULT '',
        PRIMARY KEY(school_id, student_id, term_id)
    );
    CREATE TABLE IF NOT EXISTS school_config (
        school_id INTEGER NOT NULL DEFAULT 1,
        key       TEXT NOT NULL,
        value     TEXT NOT NULL DEFAULT '',
        PRIMARY KEY(school_id, key)
    );
    CREATE TABLE IF NOT EXISTS school_subjects (
        id           SERIAL PRIMARY KEY,
        school_id    INTEGER NOT NULL DEFAULT 1,
        name         TEXT NOT NULL,
        abbreviation TEXT NOT NULL,
        sort_order   INTEGER DEFAULT 0,
        UNIQUE(school_id, name)
    );
    CREATE TABLE IF NOT EXISTS grade_config (
        id         SERIAL PRIMARY KEY,
        school_id  INTEGER NOT NULL DEFAULT 1,
        min_score  REAL NOT NULL,
        max_score  REAL NOT NULL,
        grade      TEXT NOT NULL,
        sort_order INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS announcements (
        id             SERIAL PRIMARY KEY,
        school_id      INTEGER NOT NULL DEFAULT 1,
        title          TEXT NOT NULL,
        body           TEXT NOT NULL,
        target_classes TEXT NOT NULL DEFAULT 'all',
        posted_by      TEXT NOT NULL,
        posted_at      TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS announcement_reads (
        announcement_id INTEGER NOT NULL,
        student_id      INTEGER NOT NULL,
        read_at         TIMESTAMP DEFAULT NOW(),
        PRIMARY KEY(announcement_id, student_id)
    );
    CREATE TABLE IF NOT EXISTS results_published (
        school_id INTEGER NOT NULL DEFAULT 1,
        term_id   INTEGER NOT NULL,
        published INTEGER NOT NULL DEFAULT 0,
        PRIMARY KEY(school_id, term_id)
    );
    CREATE TABLE IF NOT EXISTS superadmins (
        username TEXT PRIMARY KEY,
        password TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS platform_announcements (
        id        SERIAL PRIMARY KEY,
        title     TEXT NOT NULL,
        body      TEXT NOT NULL,
        target    TEXT NOT NULL DEFAULT 'all',
        posted_at TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS platform_announcement_reads (
        announcement_id INTEGER NOT NULL,
        school_id       INTEGER NOT NULL,
        read_at         TIMESTAMP DEFAULT NOW(),
        PRIMARY KEY(announcement_id, school_id)
    );
    CREATE TABLE IF NOT EXISTS platform_payment_config (
        id             INTEGER PRIMARY KEY DEFAULT 1,
        business_name  TEXT NOT NULL DEFAULT 'DrDemic',
        payment_number TEXT NOT NULL DEFAULT '',
        networks       TEXT NOT NULL DEFAULT 'M-Pesa,Airtel Money,Mixx,HaloPesa',
        updated_at     TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS payment_requests (
        id              SERIAL PRIMARY KEY,
        school_id       INTEGER NOT NULL,
        plan            TEXT NOT NULL,
        claimed_amount  REAL NOT NULL,
        transaction_id  TEXT NOT NULL,
        phone_used      TEXT NOT NULL,
        payment_date    DATE NOT NULL,
        note            TEXT DEFAULT '',
        status          TEXT NOT NULL DEFAULT 'pending',
        submitted_by    TEXT,
        submitted_at    TIMESTAMP DEFAULT NOW(),
        decided_by      TEXT,
        decided_at      TIMESTAMP,
        decision_note   TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS necta_grades (
    id SERIAL PRIMARY KEY,
    level TEXT NOT NULL,
    min_score REAL NOT NULL,
    max_score REAL NOT NULL,
    grade TEXT NOT NULL,
    points INTEGER NOT NULL,
    sort_order INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS necta_divisions (
        id SERIAL PRIMARY KEY,
        level TEXT NOT NULL,
        min_points INTEGER NOT NULL,
        max_points INTEGER NOT NULL,
        division TEXT NOT NULL,
        sort_order INTEGER DEFAULT 0
    );
        CREATE TABLE IF NOT EXISTS term_tests (
        id         SERIAL PRIMARY KEY,
        school_id  INTEGER NOT NULL,
        term_id    INTEGER NOT NULL,
        label      TEXT NOT NULL,
        all_classes INTEGER NOT NULL DEFAULT 1,
        created_at TIMESTAMP DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS test_classes (
        test_id  INTEGER NOT NULL,
        class_id INTEGER NOT NULL,
        PRIMARY KEY(test_id, class_id)
    );
    CREATE TABLE IF NOT EXISTS test_scores (
        id         SERIAL PRIMARY KEY,
        school_id  INTEGER NOT NULL,
        student_id INTEGER NOT NULL,
        subject    TEXT NOT NULL,
        test_id    INTEGER NOT NULL,
        score      REAL NOT NULL,
        entered_by TEXT,
        term_id    INTEGER NOT NULL,
        UNIQUE(school_id, student_id, subject, test_id, term_id)
    );
    CREATE TABLE IF NOT EXISTS published_assessments (
        school_id  INTEGER NOT NULL,
        term_id    INTEGER NOT NULL,
        assess_key TEXT NOT NULL,
        published  INTEGER DEFAULT 0,
        PRIMARY KEY(school_id, term_id, assess_key)
    );
    """)
    

    # Migrations - add missing columns to existing tables
    migrations = [
        "ALTER TABLE schools ADD COLUMN IF NOT EXISTS registered_at TIMESTAMP DEFAULT NOW()",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS is_class_teacher INTEGER DEFAULT 0",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS class_id INTEGER DEFAULT NULL",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS stream_id INTEGER DEFAULT NULL",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS must_change_password INTEGER DEFAULT 0",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS student_id INTEGER DEFAULT NULL",
        "ALTER TABLE classes ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE streams ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS stream_id INTEGER DEFAULT NULL",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS phone_number TEXT DEFAULT NULL",
        "ALTER TABLE subject_assignments ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE terms ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE ca_scores ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE exam_scores ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE remarks ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE school_config ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE school_subjects ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE grade_config ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE announcements ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE results_published ADD COLUMN IF NOT EXISTS school_id INTEGER DEFAULT 1",
        "ALTER TABLE schools ADD COLUMN IF NOT EXISTS subscription_exempt INTEGER DEFAULT 0",
        "ALTER TABLE schools ADD COLUMN IF NOT EXISTS subscription_status TEXT DEFAULT 'inactive'",
        "ALTER TABLE schools ADD COLUMN IF NOT EXISTS subscription_plan TEXT DEFAULT ''",
        "ALTER TABLE schools ADD COLUMN IF NOT EXISTS subscription_expires_at TIMESTAMP",
        "ALTER TABLE schools ADD COLUMN IF NOT EXISTS grading_system TEXT DEFAULT 'o_level'",
        "ALTER TABLE schools ADD COLUMN IF NOT EXISTS division_source TEXT DEFAULT 'school'",
        "ALTER TABLE school_subjects ADD COLUMN IF NOT EXISTS is_principal INTEGER DEFAULT 0",
        "ALTER TABLE grade_config ADD COLUMN IF NOT EXISTS points INTEGER",
        "ALTER TABLE school_subjects ADD COLUMN IF NOT EXISTS is_noncredit INTEGER DEFAULT 0",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS school_student_no INTEGER",
        "ALTER TABLE term_tests ADD COLUMN IF NOT EXISTS all_classes INTEGER NOT NULL DEFAULT 1",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS flag_reason TEXT DEFAULT NULL",
    ]
    
    for m in migrations:
        try: cur.execute(m)
        except Exception as e: print(f"Migration note: {e}")

    # Seed school_id=1 for all existing data
    try:
        cur.execute("INSERT INTO schools(id,school_name) VALUES(1,'Default School') ON CONFLICT DO NOTHING")
    except: pass

    seeds = [
        "UPDATE users SET school_id=1 WHERE school_id IS NULL",
        "UPDATE classes SET school_id=1 WHERE school_id IS NULL",
        "UPDATE streams SET school_id=1 WHERE school_id IS NULL",
        "UPDATE students SET school_id=1 WHERE school_id IS NULL",
        "UPDATE subject_assignments SET school_id=1 WHERE school_id IS NULL",
        "UPDATE terms SET school_id=1 WHERE school_id IS NULL",
        "UPDATE ca_scores SET school_id=1 WHERE school_id IS NULL",
        "UPDATE exam_scores SET school_id=1 WHERE school_id IS NULL",
        "UPDATE remarks SET school_id=1 WHERE school_id IS NULL",
        "UPDATE school_config SET school_id=1 WHERE school_id IS NULL",
        "UPDATE school_subjects SET school_id=1 WHERE school_id IS NULL",
        "UPDATE grade_config SET school_id=1 WHERE school_id IS NULL",
        "UPDATE announcements SET school_id=1 WHERE school_id IS NULL",
        "UPDATE results_published SET school_id=1 WHERE school_id IS NULL",
        "INSERT INTO school_config(school_id,key,value) VALUES(1,'school_name','School Name') ON CONFLICT DO NOTHING",
        "INSERT INTO school_config(school_id,key,value) VALUES(1,'registration_complete','0') ON CONFLICT DO NOTHING",
        "INSERT INTO school_config(school_id,key,value) VALUES(1,'phone','') ON CONFLICT DO NOTHING",
        "INSERT INTO school_config(school_id,key,value) VALUES(1,'email','') ON CONFLICT DO NOTHING",
        "INSERT INTO school_config(school_id,key,value) VALUES(1,'motto','') ON CONFLICT DO NOTHING",
        "INSERT INTO school_config(school_id,key,value) VALUES(1,'logo_path','') ON CONFLICT DO NOTHING",
        "INSERT INTO school_config(school_id,key,value) VALUES(1,'admin_phone','') ON CONFLICT DO NOTHING",
    ]
    for s in seeds:
        try: cur.execute(s)
        except Exception as e: print(f"Seed note: {e}")

    # Repair announcements from before the target-classes default-selection
    # fix — these were saved with an empty target_classes and were invisible
    # to every parent, regardless of what class they were meant for.
    try:
        cur.execute("UPDATE announcements SET target_classes='all' WHERE target_classes IS NULL OR TRIM(target_classes)=''")
    except Exception as e:
        print(f"announcement target_classes repair note: {e}")

# Subscription grandfathering — runs exactly once, on the very next deploy.
# Every school that exists at that moment (your demo schools included) gets
# exempt=1 forever. Any school registered after this point gets exempt=0
# by column default and must subscribe.
    try:
        cur.execute("CREATE TABLE IF NOT EXISTS _subscription_migration (id INTEGER PRIMARY KEY, applied INTEGER DEFAULT 0)")
        cur.execute("INSERT INTO _subscription_migration(id,applied) VALUES(1,0) ON CONFLICT(id) DO NOTHING")
        cur.execute("SELECT applied FROM _subscription_migration WHERE id=1")
        if cur.fetchone()[0] == 0:
            cur.execute("UPDATE schools SET subscription_exempt=1")
            cur.execute("UPDATE _subscription_migration SET applied=1 WHERE id=1")
            print("Subscription migration: grandfathered all existing schools.")
    except Exception as e:
        print(f"Subscription migration note: {e}")

    try:
        cur.execute("SELECT COUNT(*) FROM necta_grades")
        if cur.fetchone()[0] == 0:
            o_level = [(80,100,'A',1),(70,79,'B',2),(60,69,'C',3),(50,59,'D',4),(0,49,'F',5)]
            a_level = [(80,100,'A',1),(70,79,'B',2),(60,69,'C',3),(50,59,'D',4),
                    (40,49,'E',5),(35,39,'S',6),(0,34,'F',7)]
            for i,(lo,hi,g,p) in enumerate(o_level):
                cur.execute("INSERT INTO necta_grades(level,min_score,max_score,grade,points,sort_order) VALUES('o_level',%s,%s,%s,%s,%s)",(lo,hi,g,p,i))
            for i,(lo,hi,g,p) in enumerate(a_level):
                cur.execute("INSERT INTO necta_grades(level,min_score,max_score,grade,points,sort_order) VALUES('a_level',%s,%s,%s,%s,%s)",(lo,hi,g,p,i))
            o_div = [(7,17,'I'),(18,21,'II'),(22,25,'III'),(26,33,'IV'),(34,50,'0')]
            a_div = [(3,9,'I'),(10,12,'II'),(13,17,'III'),(18,19,'IV'),(20,21,'0')]
            for i,(lo,hi,d) in enumerate(o_div):
                cur.execute("INSERT INTO necta_divisions(level,min_points,max_points,division,sort_order) VALUES('o_level',%s,%s,%s,%s)",(lo,hi,d,i))
            for i,(lo,hi,d) in enumerate(a_div):
                cur.execute("INSERT INTO necta_divisions(level,min_points,max_points,division,sort_order) VALUES('a_level',%s,%s,%s,%s)",(lo,hi,d,i))
    except Exception as e:
        print(f"necta seed note: {e}")
    try:
        cur.execute("""UPDATE students s
                       SET school_student_no = sub.rn
                       FROM (SELECT id, ROW_NUMBER() OVER (PARTITION BY school_id ORDER BY id) AS rn
                             FROM students WHERE school_student_no IS NULL) sub
                       WHERE s.id = sub.id""")
    except Exception as e:
        print(f"school_student_no backfill note: {e}")
    try:
        cur.execute("ALTER TABLE schools ADD COLUMN IF NOT EXISTS reg_code TEXT")
        cur.execute("SELECT id FROM schools WHERE reg_code IS NULL OR reg_code=''")
        for (missing_id,) in cur.fetchall():
            cur.execute("UPDATE schools SET reg_code=%s WHERE id=%s", (f"school-{missing_id}", missing_id))
        cur.execute("SELECT 1 FROM pg_indexes WHERE indexname='idx_schools_reg_code_ci'")
        if not cur.fetchone():
            cur.execute("CREATE UNIQUE INDEX idx_schools_reg_code_ci ON schools (LOWER(reg_code))")
    except Exception as e:
        print(f"reg_code migration note: {e}")

    try:
        cur.execute("INSERT INTO platform_payment_config(id) VALUES(1) ON CONFLICT DO NOTHING")
        # Partial unique index: blocks reuse of a transaction ID that's currently
        # pending or already funded an approval, but frees it up again if a
        # request is rejected or cancelled — so a school that mistyped some
        # OTHER field can resubmit with the same (real) transaction ID.
        cur.execute("SELECT 1 FROM pg_indexes WHERE indexname='idx_payment_requests_txn_ci'")
        if not cur.fetchone():
            cur.execute("""CREATE UNIQUE INDEX idx_payment_requests_txn_ci
                           ON payment_requests (LOWER(transaction_id))
                           WHERE status IN ('pending','approved')""")
    except Exception as e:
        print(f"payment_requests migration note: {e}")

    # Superadmin from env
    sa_user = os.environ.get("SUPERADMIN_USERNAME","")
    

    # Superadmin from env
    sa_user = os.environ.get("SUPERADMIN_USERNAME","")
    sa_pass = os.environ.get("SUPERADMIN_PASSWORD","")
    if sa_user and sa_pass:
        try:
            cur.execute("SELECT username FROM superadmins WHERE username=%s", (sa_user,))
            if not cur.fetchone():
                cur.execute("INSERT INTO superadmins(username,password) VALUES(%s,%s)",
                            (sa_user, hash_password(sa_pass)))
                print(f"Superadmin '{sa_user}' created.")
        except Exception as e: print(f"Superadmin note: {e}")

    # Migration safety net: older deployments may have a `users` table created
    # before PRIMARY KEY(username, school_id) was added. Several routes rely on
    # ON CONFLICT(username, school_id), which requires a real unique constraint
    # on exactly that pair — without it, every such insert fails with
    # "no unique or exclusion constraint matching the ON CONFLICT specification".
    try:
        cur.execute("""
            SELECT 1 FROM pg_constraint
            WHERE conrelid = 'users'::regclass
              AND contype IN ('p','u')
              AND conkey = (
                  SELECT array_agg(attnum ORDER BY attnum)
                  FROM pg_attribute
                  WHERE attrelid = 'users'::regclass
                    AND attname IN ('username','school_id')
              )
        """)
        if not cur.fetchone():
            cur.execute("ALTER TABLE users ADD CONSTRAINT users_username_school_id_key UNIQUE (username, school_id)")
            print("Migration: added missing UNIQUE(username, school_id) constraint on users.")
    except Exception as e:
        print(f"Users unique-constraint check note: {e}")

    # Performance: these columns are filtered/joined on every students/marks/import
    # request. Without indexes, schools with thousands of students (e.g. bulk Excel
    # imports) see full table scans on every lookup — this gets slower as the school grows.
    try:
        cur.execute("CREATE INDEX IF NOT EXISTS idx_students_school ON students(school_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_students_school_phone ON students(school_id, phone_number)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_students_class ON students(class_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_students_stream ON students(stream_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_classes_school ON classes(school_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_streams_class ON streams(class_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_users_school_role ON users(school_id, role)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_exam_scores_student ON exam_scores(school_id, student_id)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_ca_scores_student ON ca_scores(school_id, student_id)")
    except Exception as e:
        print(f"Index creation note: {e}")

    con.commit(); cur.close(); con.close()
    print("DB ready (multi-tenant).")

    try:
            cur.execute("INSERT INTO platform_payment_config(id) VALUES(1) ON CONFLICT DO NOTHING")
            # Partial unique index: blocks reuse of a transaction ID that's currently
            # pending or already funded an approval, but frees it up again if a
            # request is rejected or cancelled — so a school that mistyped some
            # OTHER field can resubmit with the same (real) transaction ID.
            cur.execute("SELECT 1 FROM pg_indexes WHERE indexname='idx_payment_requests_txn_ci'")
            if not cur.fetchone():
                cur.execute("""CREATE UNIQUE INDEX idx_payment_requests_txn_ci
                               ON payment_requests (LOWER(transaction_id))
                               WHERE status IN ('pending','approved')""")
    except Exception as e:
        print(f"payment_requests migration note: {e}")


# ══════════════════════════════════════════════════════════════
# API ROUTES
# ══════════════════════════════════════════════════════════════

# ── SCHOOL REGISTRATION ───────────────────────────────────────
@app.route("/api/register/school", methods=["POST"])
def api_register_school():
    data         = request.form
    school_name  = data.get("school_name","").strip()
    admin_user   = data.get("admin_username","").strip()
    admin_pass   = data.get("admin_password","").strip()
    phone        = data.get("phone","").strip()
    email        = data.get("email","").strip()
    admin_phone  = data.get("admin_phone","").strip()
    motto        = data.get("motto","").strip()
    reg_code     = data.get("reg_code","").strip()
    if not school_name: return jsonify({"ok":False,"error":"School name required"}), 400
    if not admin_user or not admin_pass: return jsonify({"ok":False,"error":"Admin username and password required"}), 400
    if reg_code:
        if not valid_reg_code(reg_code):
            return jsonify({"ok":False,"error":"Registration code must be 3-32 characters: letters, numbers, underscore or hyphen only"}), 400
        if get_school_id_by_reg_code(reg_code):
            return jsonify({"ok":False,"error":f"Registration code '{reg_code}' is already taken by another school. Please choose a different one."}), 409
    else:
        reg_code = generate_unique_reg_code(school_name)
    logo_b64 = ""; logo_mime = ""
    if "logo" in request.files:
        f = request.files["logo"]
        if f and f.filename:
            ext = f.filename.rsplit(".",1)[-1].lower() if "." in f.filename else ""
            if ext not in ALLOWED_LOGO_EXT: return jsonify({"ok":False,"error":"Logo must be an image"}), 400
            raw = f.read()
            if len(raw) > 2*1024*1024:
                return jsonify({"ok":False,"error":"Logo must be smaller than 2MB"}), 400
            logo_mime = _mime_for_ext(ext)
            logo_b64  = base64.b64encode(raw).decode("ascii")
    try:
        classes_data  = json.loads(data.get("classes","[]"))
        subjects_data = json.loads(data.get("subjects","[]"))
        grades_data   = json.loads(data.get("grades","[]"))
    except Exception as e:
        return jsonify({"ok":False,"error":f"Invalid JSON: {e}"}), 400
    if not subjects_data: return jsonify({"ok":False,"error":"At least one subject required"}), 400
    if not grades_data:   return jsonify({"ok":False,"error":"At least one grade rule required"}), 400
    con = get_db(); cur = con.cursor()
    try:
        cur.execute("INSERT INTO schools(school_name,reg_code) VALUES(%s,%s) RETURNING id", (school_name, reg_code))
        school_id = cur.fetchone()[0]
        cur.execute("INSERT INTO users(username,password,role,school_id) VALUES(%s,%s,'admin',%s)",
                    (admin_user, hash_password(admin_pass), school_id))
        logo_path = f"api/logo/{school_id}" if logo_b64 else ""
        cfg = {"school_name":school_name,"phone":phone,"email":email,"admin_phone":admin_phone,
               "motto":motto,"logo_path":logo_path,"registration_complete":"1"}
        if logo_b64:
            cfg["logo_data"] = logo_b64
            cfg["logo_mime"] = logo_mime
        for k,v in cfg.items():
            cur.execute("INSERT INTO school_config(school_id,key,value) VALUES(%s,%s,%s) ON CONFLICT(school_id,key) DO UPDATE SET value=EXCLUDED.value",
                        (school_id, k, v))
        for i,s in enumerate(subjects_data):
            name = s.get("name","").strip().lower(); ab = s.get("abbreviation","").strip().upper()
            if name:
                cur.execute("INSERT INTO school_subjects(school_id,name,abbreviation,sort_order) VALUES(%s,%s,%s,%s) ON CONFLICT(school_id,name) DO UPDATE SET abbreviation=EXCLUDED.abbreviation",
                            (school_id, name, ab, i))
        for i,g in enumerate(grades_data):
            cur.execute("INSERT INTO grade_config(school_id,min_score,max_score,grade,sort_order) VALUES(%s,%s,%s,%s,%s)",
                        (school_id, float(g["min_score"]), float(g["max_score"]), str(g["grade"]).strip(), i))
        for cls in classes_data:
            cname = cls.get("name","").strip()
            if not cname: continue
            cur.execute("INSERT INTO classes(school_id,class_name) VALUES(%s,%s) ON CONFLICT(school_id,class_name) DO NOTHING RETURNING id",
                        (school_id, cname))
            row = cur.fetchone()
            if not row:
                cur.execute("SELECT id FROM classes WHERE school_id=%s AND class_name=%s", (school_id, cname))
                row = cur.fetchone()
            cid = row[0]
            for sname in cls.get("streams",[]):
                sname = sname.strip()
                if sname:
                    cur.execute("INSERT INTO streams(school_id,class_id,stream_name) VALUES(%s,%s,%s) ON CONFLICT(class_id,stream_name) DO NOTHING",
                                (school_id, cid, sname))
        con.commit()
    except Exception as e:
        con.rollback(); cur.close(); con.close()
        return jsonify({"ok":False,"error":str(e)}), 500
    cur.close(); con.close()
    return jsonify({"ok":True,"school_id":school_id})

# ── AUTH ──────────────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def api_login():
    d = request.json or {}
    reg_code = (d.get("reg_code") or "").strip()
    u, p = d.get("username","").strip(), d.get("password","")
    if not reg_code: return jsonify({"ok":False,"error":"Enter your school's registration code"}), 400
    if not u or not p: return jsonify({"ok":False,"error":"Enter username and password"}), 400
    school_id = get_school_id_by_reg_code(reg_code)
    if not school_id:
        return jsonify({"ok":False,"error":"School registration code not recognized"}), 401
    # The registration code pins the school unambiguously, so this lookup is
    # always scoped to exactly one school — a username/password that matches
    # a *different* school can never authenticate here, by construction.
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT * FROM users WHERE username=%s AND school_id=%s", (u, school_id))
    row = cur.fetchone(); cols = [x[0] for x in cur.description] if cur.description else []
    cur.close(); con.close()
    if not row: return jsonify({"ok":False,"error":"Invalid registration code, username or password"}), 401
    user = dict(zip(cols, row))
    if not verify_password(p, user["password"]):
        return jsonify({"ok":False,"error":"Invalid registration code, username or password"}), 401
    token = issue_token(user["username"], school_id, user["role"], user.get("student_id"),
                         bool(user.get("is_class_teacher", 0)), user.get("class_id"), user.get("stream_id"))
    return jsonify({"ok":True,"token":token,"user":{
        "username":             user["username"],
        "role":                 user["role"],
        "school_id":            school_id,
        "is_class_teacher":     bool(user.get("is_class_teacher",0)),
        "class_id":             user.get("class_id"),
        "stream_id":            user.get("stream_id"),
        "must_change_password": bool(user.get("must_change_password",0)),
        "student_id":           user.get("student_id"),
    },"registration_complete": is_registration_complete(school_id)})

@app.route("/api/school/info", methods=["GET"])
def api_school_info():
    reg_code = request.args.get("reg_code")
    if reg_code:
        school_id = get_school_id_by_reg_code(reg_code)
        if not school_id: return jsonify({})
    else:
        school_id = request.args.get("school_id",1,type=int)
    keys = ["school_name","phone","email","admin_phone","motto","logo_path","registration_complete"]
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT key,value FROM school_config WHERE school_id=%s AND key=ANY(%s)", (school_id, keys))
    rows = cur.fetchall(); cur.close(); con.close()
    return jsonify({r[0]:r[1] for r in rows})



@app.route("/api/setup_admin", methods=["POST"])
def api_setup_admin():
    d = request.json
    secret = d.get("secret",""); username = d.get("username","").strip(); password = d.get("password","")
    if not os.environ.get("ADMIN_SETUP_SECRET") or secret != os.environ.get("ADMIN_SETUP_SECRET"):
        return jsonify({"ok":False,"error":"Invalid setup secret"}), 403
    if not username or not password: return jsonify({"ok":False,"error":"Username and password required"}), 400
    con = get_db(); cur = con.cursor()
    cur.execute("INSERT INTO schools(id,school_name) VALUES(1,'Default School') ON CONFLICT DO NOTHING")
    cur.execute("SELECT username FROM users WHERE role='admin' AND school_id=1")
    if cur.fetchone():
        cur.close(); con.close()
        return jsonify({"ok":False,"error":"Admin already exists"}), 409
    cur.execute("INSERT INTO users(username,password,role,school_id) VALUES(%s,%s,'admin',1)",
                (username, hash_password(password)))
    for k,v in [("school_name","School Name"),("registration_complete","0"),("phone",""),
                 ("email",""),("motto",""),("logo_path",""),("admin_phone","")]:
        cur.execute("INSERT INTO school_config(school_id,key,value) VALUES(1,%s,%s) ON CONFLICT DO NOTHING",(k,v))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True,"school_id":1})

@app.route("/api/change_password", methods=["POST"])
@require_auth
def api_change_password():
    d = request.json
    username  = g.username       # can only ever change YOUR OWN password now
    school_id = g.school_id
    old_pw    = d.get("old_password","")
    new_pw    = d.get("new_password","").strip()
    if len(new_pw) < 6: return jsonify({"ok":False,"error":"Password must be at least 6 characters"}), 400
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT password FROM users WHERE username=%s AND school_id=%s", (username, school_id))
    row = cur.fetchone()
    if not row or not verify_password(old_pw, row[0]):
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Current password is incorrect"}), 401
    cur.execute("UPDATE users SET password=%s,must_change_password=0 WHERE username=%s AND school_id=%s",
                (hash_password(new_pw), username, school_id))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True})

@app.route("/uploads/logos/<filename>")
def serve_logo(filename):
    return send_from_directory(os.path.join(UPLOAD_FOLDER,"logos"), filename)


# ── SUBJECTS / GRADES ─────────────────────────────────────────
@app.route("/api/subjects", methods=["GET"])
@require_auth
def api_get_subjects():
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT id,name,abbreviation,sort_order FROM school_subjects WHERE school_id=%s ORDER BY sort_order,name",(sid,))
    rows = to_dicts(cur.fetchall(), cur); cur.close(); con.close()
    if rows: return jsonify(rows)
    return jsonify([{"id":i,"name":n,"abbreviation":_FALLBACK_ABBR.get(n,n[:4].upper()),"sort_order":i} for i,n in enumerate(_FALLBACK_SUBJECTS)])

@app.route("/api/subjects", methods=["POST"])
@require_auth
@require_role("admin")
def api_save_subjects():
    sid = g.school_id; subjects = request.json.get("subjects",[])
    con = get_db(); cur = con.cursor()
    cur.execute("DELETE FROM school_subjects WHERE school_id=%s",(sid,))
    for i,s in enumerate(subjects):
        name=s.get("name","").strip().lower(); ab=s.get("abbreviation","").strip().upper()
        if name:
            cur.execute("INSERT INTO school_subjects(school_id,name,abbreviation,sort_order) VALUES(%s,%s,%s,%s)",(sid,name,ab,i))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True})

@app.route("/api/grades", methods=["GET"])
@require_auth
def api_get_grades():
    return jsonify(get_grade_rules(g.school_id))

def get_school_grading_settings(school_id):
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT grading_system, division_source FROM schools WHERE id=%s",(school_id,))
    row=cur.fetchone(); cur.close(); con.close()
    if not row: return {"grading_system":"o_level","division_source":"school"}
    return {"grading_system":row[0] or "o_level","division_source":row[1] or "school"}

def get_necta_grades(level):
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT min_score,max_score,grade,points FROM necta_grades WHERE level=%s ORDER BY min_score DESC",(level,))
    rows=cur.fetchall(); cur.close(); con.close()
    return [{"min_score":r[0],"max_score":r[1],"grade":r[2],"points":r[3]} for r in rows]

def get_necta_divisions(level):
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT min_points,max_points,division FROM necta_divisions WHERE level=%s ORDER BY min_points",(level,))
    rows=cur.fetchall(); cur.close(); con.close()
    return [{"min_points":r[0],"max_points":r[1],"division":r[2]} for r in rows]

def get_grade_points_rules(school_id):
    settings = get_school_grading_settings(school_id)
    if settings["division_source"]=="necta":
        return get_necta_grades(settings["grading_system"])
    return get_grade_rules(school_id)  # now includes 'points' if admin set them

def grade_and_points_for_score(rules, score):
    if score is None: return "-", None
    for r in sorted(rules, key=lambda r:-r["min_score"]):
        if score >= r["min_score"]:
            return r["grade"], r.get("points")
    return "F", None

def get_principal_subjects(school_id):
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT name FROM school_subjects WHERE school_id=%s AND is_principal=1 ORDER BY sort_order",(school_id,))
    rows=cur.fetchall(); cur.close(); con.close()
    return [r[0] for r in rows]

def get_noncredit_subjects(school_id):
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT name FROM school_subjects WHERE school_id=%s AND is_noncredit=1",(school_id,))
    rows=cur.fetchall(); cur.close(); con.close()
    return [r[0] for r in rows]

def compute_division_from_finals(school_id, finals, grading_system=None, division_source=None, noncredit_override=None):
    """finals: {subject: final_score_or_None}. Returns (total_points, division_label).
    grading_system/division_source/noncredit_override let a caller (e.g. a one-off
    grade score sheet) compute division under different settings than what's saved
    in Config, without touching the school's saved settings.

    A-level uses only the student's BEST 3 credited subjects (matches how the
    NECTA division bands are scaled). O-level uses the best 7. If a student
    doesn't have enough credited subjects with marks, "INC" (incomplete) is
    returned instead of silently showing blank. If the grade configuration
    itself is missing points for a matched grade, "ERR" is returned instead
    of silently showing blank, so the admin knows to check Grading System."""
    settings = get_school_grading_settings(school_id)
    level = grading_system or settings["grading_system"]
    div_source = division_source or settings["division_source"]
    rules = get_necta_grades(level) if div_source=="necta" else get_grade_rules(school_id)
    noncredit = set(noncredit_override) if noncredit_override is not None else set(get_noncredit_subjects(school_id))

    if level == "a_level":
        use_subjects = get_principal_subjects(school_id)
        if not use_subjects:
            use_subjects = list(finals.keys())  # fallback if admin hasn't set principals yet
        min_required = 3
    else:
        use_subjects = list(finals.keys())
        min_required = 5

    use_subjects = [s for s in use_subjects if s not in noncredit]

    scored_pairs = []
    grade_config_error = False
    for s in use_subjects:
        score = finals.get(s)
        if score is None: continue
        _, points = grade_and_points_for_score(rules, score)
        if points is None:
            grade_config_error = True
            continue
        scored_pairs.append((s, points))

    if grade_config_error:
        return "ERR", "ERR"
    if len(scored_pairs) < min_required:
        return "INC", "INC"

    scored_pairs.sort(key=lambda x: x[1])
    best_n = scored_pairs[:3] if level == "a_level" else scored_pairs[:7]
    total = sum(p for _, p in best_n)

    for d in get_necta_divisions(level):
        if d["min_points"] <= total <= d["max_points"]:
            return total, d["division"]
    return total, "0"

@app.route("/api/grades", methods=["POST"])
@require_auth
@require_role("admin")
def api_save_grades():
    sid = g.school_id; grades = request.json.get("grades",[])
    con = get_db(); cur = con.cursor()
    cur.execute("DELETE FROM grade_config WHERE school_id=%s",(sid,))
    for i,g in enumerate(grades):
        cur.execute("INSERT INTO grade_config(school_id,min_score,max_score,grade,sort_order) VALUES(%s,%s,%s,%s,%s)",
                    (sid,float(g["min_score"]),float(g["max_score"]),str(g["grade"]).strip(),i))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True})

@app.route("/api/config/grading_system", methods=["GET"])
@require_auth
def api_get_grading_system():
    sid = g.school_id
    s = get_school_grading_settings(sid)
    s["principal_subjects"] = get_principal_subjects(sid)
    s["non_credit_subjects"] = get_noncredit_subjects(sid)
    return jsonify(s)

@app.route("/api/config/grading_system", methods=["POST"])
@require_auth
@require_role("admin")
def api_set_grading_system():
    sid = g.school_id; d = request.json or {}
    grading_system = d.get("grading_system")
    division_source = d.get("division_source")
    if grading_system not in ("o_level","a_level"):
        return jsonify({"ok":False,"error":"Invalid grading system"}),400
    if division_source not in ("school","necta"):
        return jsonify({"ok":False,"error":"Invalid division source"}),400
    con=get_db(); cur=con.cursor()
    cur.execute("UPDATE schools SET grading_system=%s, division_source=%s WHERE id=%s",
                (grading_system, division_source, sid))
    con.commit(); cur.close(); con.close()
    principal = d.get("principal_subjects")
    if grading_system=="a_level" and isinstance(principal, list):
        con=get_db(); cur=con.cursor()
        cur.execute("UPDATE school_subjects SET is_principal=0 WHERE school_id=%s",(sid,))
        for name in principal:
            cur.execute("UPDATE school_subjects SET is_principal=1 WHERE school_id=%s AND name=%s",(sid,name.strip().lower()))
        con.commit(); cur.close(); con.close()
    noncredit = d.get("non_credit_subjects")
    if isinstance(noncredit, list):
        con=get_db(); cur=con.cursor()
        cur.execute("UPDATE school_subjects SET is_noncredit=0 WHERE school_id=%s",(sid,))
        for name in noncredit:
            cur.execute("UPDATE school_subjects SET is_noncredit=1 WHERE school_id=%s AND name=%s",(sid,name.strip().lower()))
        con.commit(); cur.close(); con.close()
    return jsonify({"ok":True})

# ── CLASSES ───────────────────────────────────────────────────
@app.route("/api/classes", methods=["GET"])
@require_auth
def api_get_classes():
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT * FROM classes WHERE school_id=%s ORDER BY class_name",(sid,))
    classes = to_dicts(cur.fetchall(), cur); result = []
    for c in classes:
        cur.execute("SELECT * FROM streams WHERE school_id=%s AND class_id=%s ORDER BY stream_name",(sid,c["id"]))
        result.append({**c,"streams":to_dicts(cur.fetchall(),cur)})
    cur.close(); con.close(); return jsonify(result)

@app.route("/api/classes", methods=["POST"])
@require_auth
@require_role("admin")
def api_add_class():
    sid = g.school_id; name = request.json.get("class_name","").strip()
    if not name: return jsonify({"ok":False,"error":"Class name required"}),400
    con = get_db(); cur = con.cursor()
    try:
        cur.execute("INSERT INTO classes(school_id,class_name) VALUES(%s,%s) RETURNING id",(sid,name))
        new_id = cur.fetchone()[0]; con.commit()
    except psycopg2.errors.UniqueViolation:
        con.rollback(); cur.close(); con.close(); return jsonify({"ok":False,"error":"Class already exists"}),409
    cur.close(); con.close(); return jsonify({"ok":True,"id":new_id})

@app.route("/api/classes/<int:cid>", methods=["DELETE"])
@require_auth
@require_role("admin")
def api_delete_class(cid):
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM students WHERE school_id=%s AND class_id=%s",(sid,cid))
    if cur.fetchone()[0]>0:
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Students exist in this class"}),409
    cur.execute("DELETE FROM streams WHERE school_id=%s AND class_id=%s",(sid,cid))
    cur.execute("DELETE FROM classes WHERE id=%s AND school_id=%s",(cid,sid))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/classes/<int:cid>/streams", methods=["POST"])
@require_auth
@require_role("admin")
def api_add_stream(cid):
    sid = g.school_id; name = request.json.get("stream_name","").strip()
    if not name: return jsonify({"ok":False,"error":"Stream name required"}),400
    con = get_db(); cur = con.cursor()
    try:
        cur.execute("INSERT INTO streams(school_id,class_id,stream_name) VALUES(%s,%s,%s) RETURNING id",(sid,cid,name))
        new_id = cur.fetchone()[0]; con.commit()
    except psycopg2.errors.UniqueViolation:
        con.rollback(); cur.close(); con.close(); return jsonify({"ok":False,"error":"Stream already exists"}),409
    cur.close(); con.close(); return jsonify({"ok":True,"id":new_id})

@app.route("/api/streams/<int:stream_id>", methods=["DELETE"])
@require_auth
@require_role("admin")
def api_delete_stream(stream_id):
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM students WHERE school_id=%s AND stream_id=%s",(sid,stream_id))
    if cur.fetchone()[0]>0:
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Students exist in this stream"}),409
    cur.execute("DELETE FROM streams WHERE id=%s AND school_id=%s",(stream_id,sid))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

# ── STUDENTS ──────────────────────────────────────────────────
@app.route("/api/students", methods=["GET"])
@require_auth
def api_students():
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT s.id,s.name,s.class_id,s.stream_id,c.class_name,st.stream_name,s.school_student_no,s.flag_reason
                   FROM students s JOIN classes c ON s.class_id=c.id
                   LEFT JOIN streams st ON s.stream_id=st.id
                   WHERE s.school_id=%s ORDER BY c.class_name,st.stream_name,s.name""",(sid,))
    rows = to_dicts(cur.fetchall(),cur); cur.close(); con.close()
    for r in rows: r["display_id"] = format_student_display_id(sid, r.pop("school_student_no", None))
    return jsonify(rows)

@app.route("/api/students", methods=["POST"])
@require_auth
@require_role("admin","teacher")
def api_add_student():
    sid = g.school_id; d = request.json
    name=d.get("name","").strip(); class_id=d.get("class_id"); stream_id=d.get("stream_id") or None
    phone=d.get("phone_number","").strip()
    if not name or not class_id: return jsonify({"ok":False,"error":"Name and class required"}),400
    if not phone or len(phone)<4: return jsonify({"ok":False,"error":"Parent phone required"}),400
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT id FROM classes WHERE id=%s AND school_id=%s",(class_id,sid))
    if not cur.fetchone(): cur.close(); con.close(); return jsonify({"ok":False,"error":"Invalid class"}),400
    cur.execute("SELECT id FROM students WHERE school_id=%s AND LOWER(TRIM(name))=%s AND class_id=%s "
                "AND COALESCE(stream_id,0)=%s AND phone_number=%s",
                (sid, name.lower(), class_id, stream_id or 0, phone))
    if cur.fetchone():
        cur.close(); con.close()
        return jsonify({"ok":False,"error":"A student with this name, class, stream and phone already exists"}),409
    cur.execute("""INSERT INTO students(school_id,name,class_id,stream_id,phone_number,school_student_no)
                   VALUES(%s,%s,%s,%s,%s,(SELECT COALESCE(MAX(school_student_no),0)+1 FROM students WHERE school_id=%s))
                   RETURNING id""",
                (sid,name,class_id,stream_id,phone,sid))
    student_id = cur.fetchone()[0]; con.commit(); cur.close(); con.close()
    username, temp_pw = _gen_parent_creds(sid, name, phone, student_id)
    con = get_db(); cur = con.cursor()
    cur.execute("INSERT INTO users(username,password,role,school_id,must_change_password,student_id) VALUES(%s,%s,'parent',%s,1,%s) ON CONFLICT(username,school_id) DO NOTHING",
                (username, hash_password(temp_pw), sid, student_id))
    created = cur.rowcount > 0
    con.commit(); cur.close(); con.close()
    if created:
        return jsonify({"ok":True,"parent_username":username,"temp_password":temp_pw})
    return jsonify({"ok":True,"parent_username":None,"temp_password":None,
                    "warning":f"Student added, but username '{username}' is already taken by another parent account "
                              f"(same name as an existing student). No new login was created — that student will "
                              f"need a different name on file, or share the existing '{username}' login."})

def _gen_parent_creds(school_id, student_name, phone_number, student_id):
    username = student_name.strip().lower().replace(" ","_")
    last4 = phone_number.strip()[-4:]
    # Normally the password is just the last 4 digits of the parent's phone.
    # Only append "-{student_id}" when another student already shares this
    # exact username with a DIFFERENT phone number whose last 4 digits happen
    # to collide with this one.
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT s.phone_number FROM users u JOIN students s ON u.student_id = s.id
                   WHERE u.username=%s AND u.school_id=%s AND u.role='parent'""",
                (username, school_id))
    existing_phones = [r[0] or "" for r in cur.fetchall()]
    cur.close(); con.close()
    needs_suffix = any(
        ph.strip() != phone_number.strip() and ph.strip()[-4:] == last4
        for ph in existing_phones
    )
    temp_pw = f"{last4}-{student_id}" if needs_suffix else last4
    return username, temp_pw

@app.route("/api/students/resolve_display_id", methods=["GET"])
@require_auth
def api_resolve_display_id():
    sid = g.school_id
    no = request.args.get("no")
    if not no: return jsonify({"ok":False,"error":"Missing student number"}),400
    try: no = int(no)
    except ValueError: return jsonify({"ok":False,"error":"Invalid student ID"}),400
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT id FROM students WHERE school_id=%s AND school_student_no=%s",(sid,no))
    row=cur.fetchone(); cur.close(); con.close()
    if not row: return jsonify({"ok":False,"error":"Student not found"}),404
    return jsonify({"ok":True,"id":row[0]})

@app.route("/api/students/bulk_delete", methods=["POST"])
@require_auth
def api_bulk_delete_students():
    sid = g.school_id; ids = request.json.get("ids", [])
    try: ids = [int(i) for i in ids]
    except (TypeError, ValueError): return jsonify({"ok":False,"error":"Invalid student IDs"}),400
    if not ids: return jsonify({"ok":False,"error":"No students selected"}),400
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT id FROM students WHERE school_id=%s AND id=ANY(%s)",(sid,ids))
    valid_ids = [r[0] for r in cur.fetchall()]
    if not valid_ids: cur.close(); con.close(); return jsonify({"ok":False,"error":"No matching students found"}),404
    cur.execute("DELETE FROM announcement_reads WHERE student_id=ANY(%s)",(valid_ids,))
    cur.execute("DELETE FROM remarks WHERE school_id=%s AND student_id=ANY(%s)",(sid,valid_ids))
    cur.execute("DELETE FROM ca_scores WHERE school_id=%s AND student_id=ANY(%s)",(sid,valid_ids))
    cur.execute("DELETE FROM exam_scores WHERE school_id=%s AND student_id=ANY(%s)",(sid,valid_ids))
    cur.execute("DELETE FROM users WHERE school_id=%s AND student_id=ANY(%s) AND role='parent'",(sid,valid_ids))
    cur.execute("DELETE FROM students WHERE school_id=%s AND id=ANY(%s)",(sid,valid_ids))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True,"deleted":len(valid_ids)})

@app.route("/api/students/<int:student_id>", methods=["DELETE"])
@require_auth
def api_delete_student(student_id):
    sid = g.school_id;
    con = get_db(); cur = con.cursor()
    cur.execute("DELETE FROM announcement_reads WHERE student_id=%s",(student_id,))
    cur.execute("DELETE FROM remarks WHERE school_id=%s AND student_id=%s",(sid,student_id))
    cur.execute("DELETE FROM ca_scores WHERE school_id=%s AND student_id=%s",(sid,student_id))
    cur.execute("DELETE FROM exam_scores WHERE school_id=%s AND student_id=%s",(sid,student_id))
    cur.execute("DELETE FROM users WHERE school_id=%s AND student_id=%s AND role='parent'",(sid,student_id))
    cur.execute("DELETE FROM students WHERE id=%s AND school_id=%s",(student_id,sid))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/students/<int:student_id>", methods=["PATCH"])
@require_auth
@require_role("admin")
def api_update_student(student_id):
    """Edits name/class/stream/phone. If the change touches name or phone AND
    the parent hasn't logged in yet (must_change_password=1), regenerate their
    username/password to match — no point mailing someone credentials for a
    kid whose name just changed. If the parent already logged in, we leave
    their credentials alone; changing a login someone's already using is how
    you get a very confused parent and a support ticket."""
    sid = g.school_id; d = request.json or {}
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT name,class_id,stream_id,phone_number FROM students WHERE id=%s AND school_id=%s",(student_id,sid))
    row = cur.fetchone()
    if not row:
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Student not found"}),404
    old_name, old_class_id, old_stream_id, old_phone = row

    name = (d.get("name") if d.get("name") is not None else old_name).strip()
    try:
        class_id = int(d.get("class_id", old_class_id))
    except (TypeError, ValueError):
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Invalid class"}),400
    sd = d.get("stream_id", old_stream_id)
    stream_id = int(sd) if sd else None
    phone = (d.get("phone_number") if d.get("phone_number") is not None else (old_phone or "")).strip()

    if not name:
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Name required"}),400
    cur.execute("SELECT id FROM classes WHERE id=%s AND school_id=%s",(class_id,sid))
    if not cur.fetchone():
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Invalid class"}),400

    cur.execute("""SELECT id FROM students WHERE school_id=%s AND LOWER(TRIM(name))=%s AND class_id=%s
                   AND COALESCE(stream_id,0)=%s AND phone_number=%s AND id!=%s""",
                (sid, name.lower(), class_id, stream_id or 0, phone, student_id))
    if cur.fetchone():
        cur.close(); con.close()
        return jsonify({"ok":False,"error":"Another student with this name, class, stream and phone already exists"}),409

    name_changed  = name.lower() != (old_name or "").strip().lower()
    phone_changed = phone != (old_phone or "").strip()

    cur.execute("""SELECT username, must_change_password FROM users
                   WHERE school_id=%s AND student_id=%s AND role='parent'""",(sid, student_id))
    parent_row = cur.fetchone()

    new_username = new_password = cred_note = None

    if parent_row and (name_changed or phone_changed):
        old_username, must_change = parent_row
        if must_change == 1 and phone:
            gen_username = name.lower().replace(" ","_")
            last4 = phone[-4:]
            cur.execute("""SELECT s.phone_number FROM users u JOIN students s ON u.student_id=s.id
                           WHERE u.username=%s AND u.school_id=%s AND u.role='parent' AND u.student_id!=%s""",
                        (gen_username, sid, student_id))
            other_phones = [r[0] or "" for r in cur.fetchall()]
            needs_suffix = any(ph.strip()!=phone and ph.strip()[-4:]==last4 for ph in other_phones)
            new_password = f"{last4}-{student_id}" if needs_suffix else last4
            new_username = gen_username
            try:
                cur.execute("""UPDATE users SET username=%s, password=%s, must_change_password=1
                               WHERE username=%s AND school_id=%s AND student_id=%s""",
                            (new_username, hash_password(new_password), old_username, sid, student_id))
            except psycopg2.errors.UniqueViolation:
                con.rollback(); cur.close(); con.close()
                return jsonify({"ok":False,"error":f"Can't regenerate login — username '{new_username}' is already taken"}),409
            cred_note = "Parent hadn't logged in yet — credentials were regenerated to match."
        else:
            cred_note = "Name/phone updated. Parent already logged in once, so their existing login was left untouched."
    elif not parent_row and phone:
        gen_username, temp_pw = _gen_parent_creds(sid, name, phone, student_id)
        try:
            cur.execute("""INSERT INTO users(username,password,role,school_id,must_change_password,student_id)
                           VALUES(%s,%s,'parent',%s,1,%s)""",
                        (gen_username, hash_password(temp_pw), sid, student_id))
            new_username, new_password = gen_username, temp_pw
            cred_note = "Phone was missing before — a parent login was just created."
        except psycopg2.errors.UniqueViolation:
            con.rollback(); cur.close(); con.close()
            return jsonify({"ok":False,"error":f"Can't create login — username '{gen_username}' already taken"}),409

    flag_reason = "Missing parent phone — no parent login was created" if not phone else None

    cur.execute("""UPDATE students SET name=%s, class_id=%s, stream_id=%s, phone_number=%s, flag_reason=%s
                   WHERE id=%s AND school_id=%s""",
                (name, class_id, stream_id, phone or None, flag_reason, student_id, sid))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True,"new_username":new_username,"new_password":new_password,"note":cred_note})


# ── TEACHERS ──────────────────────────────────────────────────
@app.route("/api/teachers", methods=["GET"])
@require_auth
def api_get_teachers():
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT u.username,u.is_class_teacher,u.class_id,u.stream_id,u.must_change_password,
                          c.class_name,st.stream_name
                   FROM users u
                   LEFT JOIN classes c ON u.class_id=c.id AND c.school_id=%s
                   LEFT JOIN streams st ON u.stream_id=st.id
                   WHERE u.role='teacher' AND u.school_id=%s ORDER BY u.username""",(sid,sid))
    teachers = to_dicts(cur.fetchall(),cur); result=[]
    for t in teachers:
        cur.execute("""SELECT sa.subject,sa.class_id,sa.stream_id,c.class_name,st.stream_name
                       FROM subject_assignments sa JOIN classes c ON sa.class_id=c.id
                       LEFT JOIN streams st ON sa.stream_id=st.id
                       WHERE sa.school_id=%s AND sa.username=%s""",(sid,t["username"]))
        result.append({**t,"assignments":to_dicts(cur.fetchall(),cur)})
    cur.close(); con.close(); return jsonify(result)

@app.route("/api/teachers", methods=["POST"])
@require_auth
@require_role("admin")
def api_create_teacher():
    sid = g.school_id; d = request.json
    username=d.get("username","").strip(); password=d.get("password","")
    if not username or not password: return jsonify({"ok":False,"error":"Username and password required"}),400
    con = get_db(); cur = con.cursor()
    try:
        cur.execute("INSERT INTO users(username,password,role,school_id,must_change_password) VALUES(%s,%s,'teacher',%s,1)",
                    (username,hash_password(password),sid))
        con.commit()
    except psycopg2.errors.UniqueViolation:
        con.rollback(); cur.close(); con.close()
        return jsonify({"ok":False,"error":"Username already exists"}),409
    cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/teachers/<username>", methods=["DELETE"])
@require_auth
@require_role("admin")
def api_delete_teacher(username):
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("DELETE FROM subject_assignments WHERE school_id=%s AND username=%s",(sid,username))
    cur.execute("DELETE FROM users WHERE username=%s AND role='teacher' AND school_id=%s",(username,sid))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/teachers/<username>/class_teacher", methods=["POST"])
@require_auth
@require_role("admin")
def api_set_class_teacher(username):
    sid = g.school_id; d = request.json
    is_ct=bool(d.get("is_class_teacher",False)); class_id=d.get("class_id") or None; stream_id=d.get("stream_id") or None
    if is_ct and not class_id: return jsonify({"ok":False,"error":"Class required"}),400
    con = get_db(); cur = con.cursor()
    cur.execute("UPDATE users SET is_class_teacher=%s,class_id=%s,stream_id=%s WHERE username=%s AND school_id=%s AND role='teacher'",
                (1 if is_ct else 0, class_id if is_ct else None, stream_id if is_ct else None, username, sid))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/assign_teacher", methods=["POST"])
@require_auth
@require_role("admin")
def api_assign_teacher():
    sid = g.school_id; d = request.json
    username=d.get("username",""); subject=d.get("subject","").lower().strip()
    class_id=d.get("class_id"); stream_id=d.get("stream_id") or None
    con = get_db(); cur = con.cursor()
    try:
        cur.execute("INSERT INTO subject_assignments(school_id,username,subject,class_id,stream_id) VALUES(%s,%s,%s,%s,%s)",
                    (sid,username,subject,class_id,stream_id))
        con.commit()
    except psycopg2.errors.UniqueViolation:
        con.rollback(); cur.close(); con.close(); return jsonify({"ok":False,"error":"Already assigned"}),409
    cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/unassign_teacher", methods=["POST"])
@require_auth
@require_role("admin")
def api_unassign_teacher():
    sid = g.school_id; d = request.json
    username=d.get("username",""); subject=d.get("subject","").lower()
    class_id=d.get("class_id"); stream_id=d.get("stream_id") or None
    con = get_db(); cur = con.cursor()
    if stream_id:
        cur.execute("DELETE FROM subject_assignments WHERE school_id=%s AND username=%s AND subject=%s AND class_id=%s AND stream_id=%s",
                    (sid,username,subject,class_id,stream_id))
    else:
        cur.execute("DELETE FROM subject_assignments WHERE school_id=%s AND username=%s AND subject=%s AND class_id=%s AND stream_id IS NULL",
                    (sid,username,subject,class_id))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

# ── TERMS ──────────────────────────────────────────────────────
@app.route("/api/terms", methods=["GET"])
@require_auth
def api_get_terms():
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT * FROM terms WHERE school_id=%s ORDER BY id DESC",(sid,))
    rows = to_dicts(cur.fetchall(),cur); cur.close(); con.close(); return jsonify(rows)

@app.route("/api/terms/active", methods=["GET"])
@require_auth
def api_active_term():
    sid = g.school_id; t = get_active_term(sid)
    return jsonify({"ok":bool(t),"term":t})

@app.route("/api/terms", methods=["POST"])
@require_auth
@require_role("admin")
def api_create_term():
    sid = g.school_id; d = request.json
    label=d.get("label","").strip(); ca_count=int(d.get("ca_count",2))
    ca_weight=int(d.get("ca_weight",30)); ex_weight=int(d.get("exam_weight",70))
    if not label: return jsonify({"ok":False,"error":"Term label required"}),400
    if ca_weight+ex_weight!=100: return jsonify({"ok":False,"error":"Weights must sum to 100"}),400
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT id FROM terms WHERE school_id=%s AND status='open'",(sid,))
    if cur.fetchone(): cur.close(); con.close(); return jsonify({"ok":False,"error":"Close current term first"}),409
    cur.execute("INSERT INTO terms(school_id,label,ca_count,ca_weight,exam_weight,status) VALUES(%s,%s,%s,%s,%s,'open')",
                (sid,label,ca_count,ca_weight,ex_weight))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/terms/<int:tid>", methods=["PATCH"])
@require_auth
@require_role("admin")
def api_update_term(tid):
    sid = g.school_id; d = request.json or {}
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT status FROM terms WHERE id=%s AND school_id=%s",(tid,sid))
    row = cur.fetchone()
    if not row:
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Term not found"}),404
    if row[0]=="closed":
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Closed terms are locked and can't be edited"}),400

    fields=[]; vals=[]
    if d.get("label") is not None:
        label=d.get("label","").strip()
        if not label:
            cur.close(); con.close(); return jsonify({"ok":False,"error":"Term label required"}),400
        fields.append("label=%s"); vals.append(label)
    if d.get("ca_count") is not None:
        try: ca_count=int(d.get("ca_count"))
        except (TypeError,ValueError):
            cur.close(); con.close(); return jsonify({"ok":False,"error":"Invalid CA count"}),400
        if ca_count<1:
            cur.close(); con.close(); return jsonify({"ok":False,"error":"Number of CAs must be at least 1"}),400
        fields.append("ca_count=%s"); vals.append(ca_count)
    if d.get("ca_weight") is not None and d.get("exam_weight") is not None:
        try:
            ca_weight=int(d.get("ca_weight")); exam_weight=int(d.get("exam_weight"))
        except (TypeError,ValueError):
            cur.close(); con.close(); return jsonify({"ok":False,"error":"Invalid weights"}),400
        if ca_weight+exam_weight!=100:
            cur.close(); con.close(); return jsonify({"ok":False,"error":"Weights must sum to 100"}),400
        fields.append("ca_weight=%s"); vals.append(ca_weight)
        fields.append("exam_weight=%s"); vals.append(exam_weight)
    if not fields:
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Nothing to update"}),400

    vals += [tid, sid]
    cur.execute(f"UPDATE terms SET {','.join(fields)} WHERE id=%s AND school_id=%s", vals)
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True})

@app.route("/api/terms/<int:tid>/close", methods=["POST"])
@require_auth
@require_role("admin")
def api_close_term(tid):
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT status FROM terms WHERE id=%s AND school_id=%s",(tid,sid))
    row = cur.fetchone()
    if not row: cur.close(); con.close(); return jsonify({"ok":False,"error":"Not found"}),404
    if row[0]=="closed": cur.close(); con.close(); return jsonify({"ok":False,"error":"Already closed"}),400
    cur.execute("UPDATE terms SET status='closed' WHERE id=%s AND school_id=%s",(tid,sid))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

# ── MARKS ─────────────────────────────────────────────────────
@app.route("/api/marks/ca", methods=["POST"])
@require_auth
def api_enter_ca():
    sid = g.school_id; d = request.json
    username = g.username  # no longer trusts the body
    subject=d.get("subject","").lower().strip()
    class_id=int(d.get("class_id")); stream_id=d.get("stream_id") or None
    student_id=int(d.get("student_id")); ca_name=d.get("ca_name",""); score=float(d.get("score"))
    if not (0<=score<=100): return jsonify({"ok":False,"error":"Score must be 0-100"}),400
    if g.role=="teacher" and not teacher_can_access(sid,username,subject,class_id,stream_id):
        return jsonify({"ok":False,"error":"Access denied"}),403
    term = get_active_term(sid)
    if not term: return jsonify({"ok":False,"error":"No active term"}),400
    con = get_db(); cur = con.cursor()
    cur.execute("""INSERT INTO ca_scores(school_id,student_id,subject,ca_name,score,entered_by,term_id)
                   VALUES(%s,%s,%s,%s,%s,%s,%s)
                   ON CONFLICT(school_id,student_id,subject,ca_name,term_id)
                   DO UPDATE SET score=EXCLUDED.score,entered_by=EXCLUDED.entered_by""",
                (sid,student_id,subject,ca_name,score,username,term["id"]))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/marks/exam", methods=["POST"])
@require_auth
def api_enter_exam():
    sid = g.school_id; d = request.json
    username = g.username
    subject=d.get("subject","").lower().strip()
    class_id=int(d.get("class_id")); stream_id=d.get("stream_id") or None
    student_id=int(d.get("student_id")); score=float(d.get("score"))
    if not (0<=score<=100): return jsonify({"ok":False,"error":"Score must be 0-100"}),400
    if g.role=="teacher" and not teacher_can_access(sid,username,subject,class_id,stream_id):
        return jsonify({"ok":False,"error":"Access denied"}),403
    term = get_active_term(sid)
    if not term: return jsonify({"ok":False,"error":"No active term"}),400
    con = get_db(); cur = con.cursor()
    cur.execute("""INSERT INTO exam_scores(school_id,student_id,subject,score,entered_by,term_id)
                   VALUES(%s,%s,%s,%s,%s,%s)
                   ON CONFLICT(school_id,student_id,subject,term_id)
                   DO UPDATE SET score=EXCLUDED.score,entered_by=EXCLUDED.entered_by""",
                (sid,student_id,subject,score,username,term["id"]))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

# ── CONFIG ─────────────────────────────────────────────────────
@app.route("/api/config", methods=["GET"])
@require_auth
def api_config():
    sid = g.school_id; term = get_active_term(sid)
    subjects = get_subjects(sid); subj_map = get_subject_map(sid)
    info = {k: get_config_val(sid,k,"") for k in ["school_name","phone","email","admin_phone","motto","logo_path"]}
    return jsonify({"allowed_subjects":subjects,"subject_abbr":subj_map,"active_term":term,
                    "ca_count":term["ca_count"] if term else 2,"school_name":info.get("school_name","School Name"),
                    "school_info":info,"grade_rules":get_grade_rules(sid)})

@app.route("/api/config/school_name", methods=["POST"])
@require_auth
@require_role("admin")
def api_set_school_name():
    sid = g.school_id; name = request.json.get("school_name","").strip()
    if not name: return jsonify({"ok":False,"error":"Name cannot be empty"}),400
    set_config_val(sid,"school_name",name); return jsonify({"ok":True})

@app.route("/api/config/reg_code", methods=["GET"])
@require_auth
@require_role("admin")
def api_get_reg_code():
    sid = g.school_id()
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT reg_code FROM schools WHERE id=%s", (sid,))
    row = cur.fetchone(); cur.close(); con.close()
    return jsonify({"reg_code": row[0] if row else ""})

@app.route("/api/config/reg_code", methods=["POST"])
@require_auth
@require_role("admin")
def api_set_reg_code():
    sid = g.school_id
    new_code = ((request.json or {}).get("reg_code") or "").strip()
    if not valid_reg_code(new_code):
        return jsonify({"ok":False,"error":"Registration code must be 3-32 characters: letters, numbers, underscore or hyphen only"}), 400
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT id FROM schools WHERE LOWER(reg_code)=LOWER(%s) AND id!=%s", (new_code, sid))
    if cur.fetchone():
        cur.close(); con.close()
        return jsonify({"ok":False,"error":"That registration code is already taken by another school. Please choose a different one."}), 409
    cur.execute("UPDATE schools SET reg_code=%s WHERE id=%s", (new_code, sid))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True,"reg_code":new_code})

@app.route("/api/config/school_info", methods=["POST"])
@require_auth
@require_role("admin")
def api_set_school_info():
    sid = g.school_id; d = request.json
    for key in ["school_name","phone","email","admin_phone","motto"]:
        val = d.get(key)
        if val is not None: set_config_val(sid, key, val.strip())
    return jsonify({"ok":True})

# ── REPORT CARD ───────────────────────────────────────────────
@app.route("/api/report/<int:student_id>", methods=["GET"])
@require_auth
@require_role("parent","admin","class_teacher")
def api_report(student_id):
    sid = g.school_id; subjects = get_subjects(sid)
    term_id = request.args.get("term_id")
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT s.id,s.name,s.class_id,s.stream_id,c.class_name,st.stream_name,s.school_student_no
                   FROM students s JOIN classes c ON s.class_id=c.id
                   LEFT JOIN streams st ON s.stream_id=st.id
                   WHERE s.id=%s AND s.school_id=%s""",(student_id,sid))
    row = cur.fetchone(); student = to_dict(row,cur) if row else None
    cur.close(); con.close()
    if not student: return jsonify({"ok":False,"error":"Student not found"}),404
    student["display_id"] = format_student_display_id(sid, student.pop("school_student_no", None))
    term = get_term_by_id(sid,int(term_id)) if term_id else get_active_term(sid)
    if not term: return jsonify({"ok":False,"error":"No term available"}),400
    tid=term["id"]; ca_count=term["ca_count"]; ca_w=term["ca_weight"]; ex_w=term["exam_weight"]
    class_id=student["class_id"]; stream_id=student["stream_id"]

    class_rows, class_rank_map, stream_rank_map, scores_bulk = get_class_report_data(
        sid, tid, class_id, stream_id, subjects, ca_w, ex_w)
    subject_rank_maps = {subj: get_subject_rank_map(class_rows, subj) for subj in subjects}

    c_entry = class_rank_map.get(student_id)
    c_pos   = c_entry["position"] if c_entry else "-"
    c_total = len(class_rows)
    s_pos = s_total = None
    if stream_id and stream_rank_map is not None:
        s_entry = stream_rank_map.get(student_id)
        s_pos   = s_entry["position"] if s_entry else "-"
        s_total = len(stream_rank_map)

    student_finals = c_entry["finals"] if c_entry else compute_student_finals(scores_bulk, student_id, subjects, ca_w, ex_w)
    avg = c_entry["average"] if c_entry else round(compute_average_from_finals(student_finals), 2)
    student_scores = scores_bulk.get(student_id, {})
    finals_for_division = student_finals  # you already have this dict
    division_points, division = compute_division_from_finals(sid, finals_for_division)

    rows = []
    for subject in subjects:
        entry     = student_scores.get(subject, {})
        ca_map    = entry.get("ca", {})
        ca_scores = {f"CA{i}": ca_map.get(f"CA{i}") for i in range(1,ca_count+1)}
        exam_val  = entry.get("exam")
        final_val = student_finals.get(subject)
        subj_pos  = subject_rank_maps[subject].get(student_id, "-") if final_val is not None else "-"
        rows.append({"subject":subject,"ca":ca_scores,"exam":exam_val,
                     "final":round(final_val,1) if final_val is not None else None,
                     "grade":get_grade(sid,final_val) if final_val is not None else "-","position":subj_pos})
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT * FROM remarks WHERE school_id=%s AND student_id=%s AND term_id=%s",(sid,student_id,tid))
    rmk_row = cur.fetchone(); rmk = to_dict(rmk_row,cur) if rmk_row else None
    cur.close(); con.close()
    return jsonify({"ok":True,"student":student,"term":term,"rows":rows,
                    "average":avg,"grade":get_grade(sid,avg),
                    "class_position":c_pos,"class_total":c_total,
                    "stream_position":s_pos,"stream_total":s_total,
                    "class_teacher_remark":rmk["class_teacher_remark"] if rmk else "",
                    "head_remark":rmk["head_remark"] if rmk else "",
                    "ca_count":ca_count,"ca_weight":term["ca_weight"],"exam_weight":term["exam_weight"],"division":division,"division_points":division_points})

# ── REMARKS ───────────────────────────────────────────────────
@app.route("/api/remarks", methods=["POST"])
@require_auth
def api_remarks():
    sid = g.school_id; d = request.json
    username = g.username; role = g.role; is_ct=d.get("is_class_teacher",False)
    student_id=int(d.get("student_id")); remark=d.get("remark","").strip()
    term = get_active_term(sid)
    if not term: return jsonify({"ok":False,"error":"No active term"}),400
    tid=term["id"]
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT class_id FROM students WHERE id=%s AND school_id=%s",(student_id,sid))
    student = cur.fetchone()
    if not student: cur.close(); con.close(); return jsonify({"ok":False,"error":"Student not found"}),404
    if role=="admin": field="head_remark"
    elif role=="teacher" and is_ct:
        cur.execute("SELECT class_id FROM users WHERE username=%s AND school_id=%s AND is_class_teacher=1",(username,sid))
        u = cur.fetchone()
        if not u or u[0]!=student[0]:
            cur.close(); con.close(); return jsonify({"ok":False,"error":"Not your class"}),403
        field="class_teacher_remark"
    else:
        cur.close(); con.close(); return jsonify({"ok":False,"error":"Not allowed"}),403
    cur.execute(f"""INSERT INTO remarks(school_id,student_id,term_id,{field}) VALUES(%s,%s,%s,%s)
                    ON CONFLICT(school_id,student_id,term_id) DO UPDATE SET {field}=EXCLUDED.{field}""",
                (sid,student_id,tid,remark))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

# ── RANKINGS ──────────────────────────────────────────────────
@app.route("/api/ranking/subject", methods=["GET"])
@require_auth
def api_subject_ranking():
    sid=g.school_id; subject=request.args.get("subject","").lower()
    class_id=request.args.get("class_id"); stream_id=request.args.get("stream_id") or None
    assess=request.args.get("assess","exam"); term_id=request.args.get("term_id")
    if not term_id:
        term=get_active_term(sid)
        if not term: return jsonify([])
        term_id=term["id"]
    else: term_id=int(term_id)
    if stream_id: stream_id=int(stream_id)
    if class_id:  class_id=int(class_id)
    studs=get_students_in_scope(sid,class_id,stream_id)
    if not studs: return jsonify([])
    student_ids=[s["id"] for s in studs]
    con=get_db(); cur=con.cursor()
    if assess=="exam":
        cur.execute("SELECT student_id,score FROM exam_scores WHERE school_id=%s AND student_id=ANY(%s) AND subject=%s AND term_id=%s",
                    (sid,student_ids,subject,term_id))
    else:
        cur.execute("SELECT student_id,score FROM ca_scores WHERE school_id=%s AND student_id=ANY(%s) AND subject=%s AND ca_name=%s AND term_id=%s",
                    (sid,student_ids,subject,assess,term_id))
    score_map=dict(cur.fetchall()); cur.close(); con.close()
    name_map={s["id"]:s["name"] for s in studs}
    rows=[{"id":stid,"name":name_map[stid],"score":round(sc,2),"grade":get_grade(sid,sc)}
          for stid,sc in score_map.items()]
    _assign_positions(rows,"score"); return jsonify(rows)



# ── ANALYTICS ROUTES ───────────────────────────────────────────
@app.route("/api/analytics/overview", methods=["GET"])
@subscription_required
@require_auth
def api_analytics_overview():
    sid = g.school_id
    username  = request.args.get("username", "")
    role      = request.args.get("role", "")
    class_id  = request.args.get("class_id") or None
    stream_id = request.args.get("stream_id") or None
    class_id  = int(class_id) if class_id else None
    stream_id = int(stream_id) if stream_id else None

    # A class teacher can only ever see their own class/stream — enforced
    # server-side regardless of what the request asked for, since this is
    # the same trust boundary marks entry already relies on.
    if role == "teacher":
        con = get_db(); cur = con.cursor()
        cur.execute("SELECT is_class_teacher,class_id,stream_id FROM users WHERE username=%s AND school_id=%s",
                    (username, sid))
        row = cur.fetchone(); cur.close(); con.close()
        if not row or not row[0]:
            return jsonify({"ok": False, "error": "Not a class teacher"}), 403
        if row[1] is None:
            return jsonify({"ok": False, "error": "No class assigned"}), 403
        class_id, stream_id = row[1], row[2]

    subjects = get_subjects(sid)
    points, name_map = _compute_overall_series(sid, class_id, stream_id, subjects)
    result = _build_common_cards(points, name_map, sid)
    if points:
        best_subj, weak_subj = _best_weakest_subject(
            sid, class_id, stream_id, subjects, points[-1]["term_id"], points[-1]["assess"])
        result["best_subject"] = best_subj
        result["weakest_subject"] = weak_subj
    else:
        result["best_subject"] = None
        result["weakest_subject"] = None
    return jsonify({"ok": True, **result})

@app.route("/api/analytics/subject", methods=["GET"])
@subscription_required
@require_auth
def api_analytics_subject():
    sid       = g.school_id
    username  = request.args.get("username", "")
    role      = request.args.get("role", "")
    subject   = request.args.get("subject", "").lower().strip()
    class_id  = request.args.get("class_id")
    stream_id = request.args.get("stream_id") or None
    if not subject or not class_id:
        return jsonify({"ok": False, "error": "subject and class_id required"}), 400
    class_id  = int(class_id)
    stream_id = int(stream_id) if stream_id else None
    if role == "teacher" and not teacher_can_access(sid, username, subject, class_id, stream_id):
        return jsonify({"ok": False, "error": "Access denied"}), 403

    points, name_map = _compute_subject_series(sid, class_id, stream_id, subject)
    result = _build_common_cards(points, name_map, sid)
    return jsonify({"ok": True, **result})

@app.route("/api/analytics/dashboard_classes", methods=["GET"])
@subscription_required
@require_auth
@require_role("admin")
def api_analytics_dashboard_classes():
    sid = g.school_id
    subjects = get_subjects(sid)
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT id,class_name FROM classes WHERE school_id=%s ORDER BY class_name", (sid,))
    classes = to_dicts(cur.fetchall(), cur); cur.close(); con.close()
    results = []
    for c in classes:
        points, _ = _compute_overall_series(sid, c["id"], None, subjects)
        if points:
            results.append({"class_id": c["id"], "class_name": c["class_name"], "average": points[-1]["avg"]})
    if not results:
        return jsonify({"ok": True, "best": None, "weakest": None})
    best = max(results, key=lambda r: r["average"])
    weakest = min(results, key=lambda r: r["average"])
    return jsonify({"ok": True, "best": best, "weakest": weakest})

# ── SCORE SHEETS ──────────────────────────────────────────────
@app.route("/api/scoresheet", methods=["GET"])
@require_auth
def api_scoresheet():
    sid=g.school_id; subjects=get_subjects(sid)
    mode=request.args.get("mode","ca"); class_id=request.args.get("class_id")
    stream_id=request.args.get("stream_id") or None; ca_name=request.args.get("ca_name","CA1")
    term_id=request.args.get("term_id")
    sheet_type=request.args.get("sheet_type","marks")  # "marks" or "grade"
    grading_system=request.args.get("grading_system") or None
    division_source=request.args.get("division_source") or None
    noncredit_param=request.args.get("noncredit","")
    noncredit_override=[x.strip().lower() for x in noncredit_param.split(",") if x.strip()] if noncredit_param else None
    if not term_id:
        term=get_active_term(sid)
        if not term: return jsonify({"subjects":[],"results":[]})
        term_id=term["id"]
    else: term_id=int(term_id)
    if class_id:  class_id=int(class_id)
    if stream_id: stream_id=int(stream_id)

    studs=get_students_in_scope(sid,class_id,stream_id)
    if not studs:
        return jsonify({"subjects":subjects,"results":[],"sheet_type":sheet_type})
    student_ids=[s["id"] for s in studs]

    con=get_db(); cur=con.cursor()
    ca_scores={}; exam_scores={}; ca_avgs={}; term=None
    if mode=="ca":
        cur.execute("""SELECT student_id,subject,score FROM ca_scores
                       WHERE school_id=%s AND term_id=%s AND ca_name=%s AND student_id=ANY(%s)""",
                    (sid,term_id,ca_name,student_ids))
        for student_id,subject,score in cur.fetchall(): ca_scores[(student_id,subject)]=score
    elif mode=="exam":
        cur.execute("""SELECT student_id,subject,score FROM exam_scores
                       WHERE school_id=%s AND term_id=%s AND student_id=ANY(%s)""",
                    (sid,term_id,student_ids))
        for student_id,subject,score in cur.fetchall(): exam_scores[(student_id,subject)]=score
    elif mode=="test":
        test_id = int(request.args.get("test_id"))
        cur.execute("""SELECT student_id,subject,score FROM test_scores
                    WHERE school_id=%s AND term_id=%s AND test_id=%s AND student_id=ANY(%s)""",
                    (sid,term_id,test_id,student_ids))
        for student_id,subject,score in cur.fetchall(): exam_scores[(student_id,subject)]=score

    elif mode=="terminal":
        term=get_term_by_id(sid,term_id)
        if not term:
            cur.close(); con.close()
            return jsonify({"subjects":subjects,"results":[],"sheet_type":sheet_type})
        cur.execute("""SELECT student_id,subject,score FROM exam_scores
                       WHERE school_id=%s AND term_id=%s AND student_id=ANY(%s)""",
                    (sid,term_id,student_ids))
        for student_id,subject,score in cur.fetchall(): exam_scores[(student_id,subject)]=score
        cur.execute("""SELECT student_id,subject,AVG(score) FROM ca_scores
                       WHERE school_id=%s AND term_id=%s AND student_id=ANY(%s)
                       GROUP BY student_id,subject""",
                    (sid,term_id,student_ids))
        for student_id,subject,avg_score in cur.fetchall(): ca_avgs[(student_id,subject)]=float(avg_score)
    cur.close(); con.close()

    def score_for(stid, subject):
        if mode=="ca": return ca_scores.get((stid,subject))
        if mode in ("exam","test"): return exam_scores.get((stid,subject))
        if mode=="terminal":
            exam=exam_scores.get((stid,subject)); ca_avg=ca_avgs.get((stid,subject))
            if exam is not None and ca_avg is not None:
                return round((ca_avg/100)*term["ca_weight"] + (exam/100)*term["exam_weight"],1)
        return None

    if sheet_type=="grade":
        settings = get_school_grading_settings(sid)
        level = grading_system or settings["grading_system"]
        div_source = division_source or settings["division_source"]
        rules = get_necta_grades(level) if div_source=="necta" else get_grade_rules(sid)
        results=[]
        for s in studs:
            subj_scores={}; grades={}
            for subject in subjects:
                score = score_for(s["id"], subject)
                subj_scores[subject]=score
                grades[subject],_ = grade_and_points_for_score(rules, score)
            points, division = compute_division_from_finals(sid, subj_scores, grading_system, division_source, noncredit_override)
            results.append({"id":s["id"],"name":s["name"],"stream_name":s.get("stream_name"),
                            "grades":grades,"points":points,"division":division or "-"})
        return jsonify({"subjects":subjects,"results":results,"sheet_type":"grade"})

    grade_rules=get_grade_rules(sid)
    def grade_for(score):
        if score is None: return "-"
        for r in grade_rules:
            if score>=r["min_score"]: return r["grade"]
        return "F"

    results=[]
    for s in studs:
        row={"id":s["id"],"name":s["name"],"stream_name":s.get("stream_name"),"scores":{},"total":0,"count":0}
        for subject in subjects:
            score=score_for(s["id"], subject)
            row["scores"][subject]=score
            if score is not None: row["total"]+=score; row["count"]+=1
        row["average"]=round(row["total"]/row["count"],2) if row["count"] else 0
        row["grade"]=grade_for(row["average"]); results.append(row)
    _assign_positions(results,"average")
    return jsonify({"subjects":subjects,"results":results,"sheet_type":"marks"})


# ── ANNOUNCEMENTS ─────────────────────────────────────────────
@app.route("/api/announcements", methods=["GET"])
@require_auth
def api_get_announcements():
    sid=g.school_id; student_id=request.args.get("student_id")
    con=get_db(); cur=con.cursor()
    if student_id:
        stid=int(student_id)
        cur.execute("SELECT class_id FROM students WHERE id=%s AND school_id=%s",(stid,sid))
        row=cur.fetchone()
        if not row: cur.close(); con.close(); return jsonify([])
        cur.execute("SELECT class_name FROM classes WHERE id=%s AND school_id=%s",(row[0],sid))
        cls_row=cur.fetchone(); class_name=cls_row[0] if cls_row else ""
        cur.execute("""SELECT a.id,a.title,a.body,a.target_classes,a.posted_by,CAST(a.posted_at AS TEXT),
                              CASE WHEN ar.student_id IS NOT NULL THEN 1 ELSE 0 END as is_read
                       FROM announcements a
                       LEFT JOIN announcement_reads ar ON ar.announcement_id=a.id AND ar.student_id=%s
                       WHERE a.school_id=%s AND (a.target_classes='all' OR a.target_classes LIKE %s)
                       ORDER BY a.posted_at DESC""",(stid,sid,f"%{class_name}%"))
    else:
        cur.execute("SELECT id,title,body,target_classes,posted_by,CAST(posted_at AS TEXT),0 as is_read FROM announcements WHERE school_id=%s ORDER BY posted_at DESC",(sid,))
    rows=to_dicts(cur.fetchall(),cur); cur.close(); con.close(); return jsonify(rows)

@app.route("/api/announcements", methods=["POST"])
@subscription_required
@require_auth
@require_role("admin")
def api_post_announcement():
    sid=g.school_id; d=request.json
    title=d.get("title","").strip(); body=d.get("body","").strip()
    posted_by=d.get("posted_by",""); target_classes=(d.get("target_classes") or "all").strip() or "all"
    if not title or not body: return jsonify({"ok":False,"error":"Title and body required"}),400
    con=get_db(); cur=con.cursor()
    cur.execute("INSERT INTO announcements(school_id,title,body,target_classes,posted_by) VALUES(%s,%s,%s,%s,%s) RETURNING id",
                (sid,title,body,target_classes,posted_by))
    new_id=cur.fetchone()[0]; con.commit(); cur.close(); con.close()
    return jsonify({"ok":True,"id":new_id})

@app.route("/api/announcements/<int:aid>", methods=["DELETE"])
@require_auth
@require_role("admin")
def api_delete_announcement(aid):
    sid=g.school_id
    con=get_db(); cur=con.cursor()
    cur.execute("DELETE FROM announcements WHERE id=%s AND school_id=%s",(aid,sid))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/announcements/<int:aid>/read", methods=["POST"])
@require_auth
def api_mark_announcement_read(aid):
    student_id=request.json.get("student_id")
    if not student_id: return jsonify({"ok":False,"error":"student_id required"}),400
    con=get_db(); cur=con.cursor()
    cur.execute("INSERT INTO announcement_reads(announcement_id,student_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",(aid,int(student_id)))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

# ── RESULTS PUBLISHING ────────────────────────────────────────
@app.route("/api/results/status", methods=["GET"])
@require_auth
def api_results_status():
    sid=g.school_id; term_id=request.args.get("term_id")
    if not term_id:
        term=get_active_term(sid)
        if not term: return jsonify({"published":False,"term":None,"term_id":None})
        term_id=term["id"]
    else: term_id=int(term_id)
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT published FROM results_published WHERE school_id=%s AND term_id=%s",(sid,term_id))
    row=cur.fetchone(); cur.close(); con.close()
    return jsonify({"published":bool(row[0]) if row else False,"term":get_term_by_id(sid,term_id),"term_id":term_id})

@app.route("/api/results/toggle", methods=["POST"])
@require_auth
def api_toggle_results():
    sid=g.school_id; d=request.json
    term_id=d.get("term_id"); publish=bool(d.get("publish",True))
    if publish and not is_subscribed(sid):
        return jsonify({"ok": False, "error": "subscription_required",
                        "message": "Publishing results requires an active subscription."}), 402
    if not term_id: return jsonify({"ok":False,"error":"term_id required"}),400
    con=get_db(); cur=con.cursor()
    cur.execute("""INSERT INTO results_published(school_id,term_id,published) VALUES(%s,%s,%s)
                   ON CONFLICT(school_id,term_id) DO UPDATE SET published=EXCLUDED.published""",
                (sid,int(term_id),1 if publish else 0))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True,"published":publish})

@app.route("/api/results/assessments", methods=["GET"])
@require_auth
def api_list_assessments_for_publish():
    sid = g.school_id; term_id = request.args.get("term_id")
    if not term_id:
        term = get_active_term(sid)
        if not term: return jsonify({"ok":True,"assessments":[]})
        term_id = term["id"]
    else: term_id = int(term_id)
    term = get_term_by_id(sid, term_id)
    if not term: return jsonify({"ok":False,"error":"Term not found"}),404
    tests = get_term_tests(sid, term_id)
    test_map = {t["id"]: t["label"] for t in tests}
    keys = [f"CA{i}" for i in range(1, term["ca_count"]+1)] + ["exam"] + [f"test:{t['id']}" for t in tests]
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT assess_key, published FROM published_assessments WHERE school_id=%s AND term_id=%s",(sid,term_id))
    pub_map = dict(cur.fetchall()); cur.close(); con.close()
    result=[]
    for k in keys:
        label = "Final Exam" if k=="exam" else (test_map.get(int(k.split(":")[1])) if k.startswith("test:") else k)
        result.append({"assess_key":k, "label":label, "published": bool(pub_map.get(k,0))})
    return jsonify({"ok":True,"term_id":term_id,"assessments":result})

@app.route("/api/results/publish_assessments", methods=["POST"])
@require_auth
def api_publish_assessments():
    sid = g.school_id; d = request.json or {}
    term_id = d.get("term_id"); keys = d.get("assess_keys") or []; publish = bool(d.get("publish", True))
    if publish and not is_subscribed(sid):
        return jsonify({"ok":False,"error":"subscription_required","message":"Publishing results requires an active subscription."}),402
    if not term_id or not keys: return jsonify({"ok":False,"error":"term_id and assess_keys required"}),400
    con=get_db(); cur=con.cursor()
    for k in keys:
        cur.execute("""INSERT INTO published_assessments(school_id,term_id,assess_key,published) VALUES(%s,%s,%s,%s)
                       ON CONFLICT(school_id,term_id,assess_key) DO UPDATE SET published=EXCLUDED.published""",
                    (sid,int(term_id),k,1 if publish else 0))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True})

# ── PARENT PORTAL ─────────────────────────────────────────────
@app.route("/api/parent/terms", methods=["GET"])
@require_auth
def api_parent_terms():
    sid=g.school_id
    con=get_db(); cur=con.cursor()
    # A term shows up here if EITHER the legacy whole-term publish switch is
    # on, OR at least one individual assessment has been published via the
    # newer per-assessment publisher — previously only the legacy flag was
    # checked, so terms published assessment-by-assessment (the normal flow)
    # never appeared for parents.
    cur.execute("""SELECT DISTINCT t.id,t.label,t.ca_count,t.ca_weight,t.exam_weight,t.status
                   FROM terms t
                   WHERE t.school_id=%s AND (
                       EXISTS (SELECT 1 FROM results_published rp WHERE rp.school_id=t.school_id AND rp.term_id=t.id AND rp.published=1)
                       OR EXISTS (SELECT 1 FROM published_assessments pa WHERE pa.school_id=t.school_id AND pa.term_id=t.id AND pa.published=1)
                   )
                   ORDER BY t.id ASC""",(sid,))
    rows=to_dicts(cur.fetchall(),cur); cur.close(); con.close(); return jsonify(rows)

@app.route("/api/parent/results", methods=["GET"])
@require_auth
@require_role("parent")
def api_parent_results():
    sid=g.school_id; subjects=get_subjects(sid)
    student_id=request.args.get("student_id"); term_id=request.args.get("term_id"); assess=request.args.get("assess")
    if not student_id: return jsonify({"ok":False,"error":"student_id required"}),400
    if assess:
        con=get_db(); cur=con.cursor()
        cur.execute("SELECT published FROM published_assessments WHERE school_id=%s AND term_id=%s AND assess_key=%s",(sid,term_id,assess))
        prow=cur.fetchone(); cur.close(); con.close()
        if not prow or not prow[0]:
            return jsonify({"ok":False,"error":"This assessment hasn't been published yet"}),403
    else:
        con=get_db(); cur=con.cursor()
        cur.execute("SELECT published FROM results_published WHERE school_id=%s AND term_id=%s",(sid,term_id))
        row=cur.fetchone(); cur.close(); con.close()
        if not row or not row[0]: return jsonify({"ok":False,"error":"Results not yet published"}),403
    stid=int(student_id); term=get_term_by_id(sid,term_id)
    con=get_db(); cur=con.cursor()
    cur.execute("""SELECT s.id,s.name,s.class_id,s.stream_id,c.class_name,st.stream_name
                   FROM students s JOIN classes c ON s.class_id=c.id LEFT JOIN streams st ON s.stream_id=st.id
                   WHERE s.id=%s AND s.school_id=%s""",(stid,sid))
    row=cur.fetchone(); student=to_dict(row,cur) if row else None; cur.close(); con.close()
    if not student: return jsonify({"ok":False,"error":"Student not found"}),404
    ca_count=term["ca_count"]; ca_w=term["ca_weight"]; ex_w=term["exam_weight"]
    class_id=student["class_id"]; stream_id=student["stream_id"]

    class_rows, class_rank_map, stream_rank_map, scores_bulk = get_class_report_data(
        sid, term_id, class_id, stream_id, subjects, ca_w, ex_w)
    class_ids = [r["id"] for r in class_rows]
    student_scores = scores_bulk.get(stid, {})

    results=[]
    if assess:
        assess_rank_maps={}
        for subject in subjects:
            entry=student_scores.get(subject,{})
            ca_map=entry.get("ca",{})
            ca_scores={f"CA{i}": ca_map.get(f"CA{i}") for i in range(1,ca_count+1)}
            exam_val=entry.get("exam")
            score=_score_for_assess(entry, assess)
            if score is None: continue
            if subject not in assess_rank_maps:
                assess_rank_maps[subject]=get_subject_assess_rank_map(scores_bulk,class_ids,subject,assess)
            pos=assess_rank_maps[subject].get(stid,"-")
            results.append({"subject":subject,"ca":ca_scores,"exam":exam_val,"score":score,
                            "grade":get_grade(sid,score),"position":pos})
    else:
        subject_rank_maps={subj: get_subject_rank_map(class_rows, subj) for subj in subjects}
        student_finals = class_rank_map.get(stid,{}).get("finals") or compute_student_finals(scores_bulk,stid,subjects,ca_w,ex_w)
        for subject in subjects:
            entry=student_scores.get(subject,{})
            ca_map=entry.get("ca",{})
            ca_scores={f"CA{i}": ca_map.get(f"CA{i}") for i in range(1,ca_count+1)}
            exam_val=entry.get("exam")
            if not ca_map and exam_val is None: continue
            final_val=student_finals.get(subject)
            subj_pos=subject_rank_maps[subject].get(stid,"-") if final_val is not None else "-"
            results.append({"subject":subject,"ca":ca_scores,"exam":exam_val,
                            "final":round(final_val,1) if final_val is not None else None,
                            "grade":get_grade(sid,final_val) if final_val is not None else "-","position":subj_pos})

    if assess:
        # Rank by THIS assessment's own average across subjects, not the term's
        # weighted final. Finals are only computable once both CA and exam marks
        # exist, so before the exam is entered every student's final average is
        # 0 — that ties the whole class and everyone was showing up as "1st".
        stream_id_map = {r["id"]: r.get("stream_id") for r in class_rows}
        assess_scores=[]
        for cid in class_ids:
            student_data = scores_bulk.get(cid, {})
            vals=[]
            for subject in subjects:
                entry = student_data.get(subject, {})
                v = _score_for_assess(entry, assess)
                if v is not None: vals.append(v)
            if vals: assess_scores.append({"id":cid,"score":sum(vals)/len(vals)})
        _assign_positions(assess_scores,"score")
        assess_pos_map={r["id"]:r["position"] for r in assess_scores}
        c_pos = assess_pos_map.get(stid,"-")
        c_total = len(assess_scores)
        s_pos=s_total=None
        if stream_id:
            stream_scores=[dict(r) for r in assess_scores if stream_id_map.get(r["id"])==stream_id]
            _assign_positions(stream_scores,"score")
            stream_pos_map={r["id"]:r["position"] for r in stream_scores}
            s_pos = stream_pos_map.get(stid,"-")
            s_total = len(stream_scores)
        if results:
            scores_only=[r["score"] for r in results if r.get("score") is not None]
            avg=round(sum(scores_only)/len(scores_only),2) if scores_only else 0
        else:
            avg = 0
    else:
        c_entry=class_rank_map.get(stid)
        c_pos=c_entry["position"] if c_entry else "-"
        c_total=len(class_rows)
        s_pos=s_total=None
        if stream_id and stream_rank_map is not None:
            s_entry=stream_rank_map.get(stid)
            s_pos=s_entry["position"] if s_entry else "-"
            s_total=len(stream_rank_map)
        avg = c_entry["average"] if c_entry else round(compute_average_from_finals(compute_student_finals(scores_bulk,stid,subjects,ca_w,ex_w)),2)

    if assess:
        subj_scores_div = {r["subject"]: r["score"] for r in results if r.get("score") is not None}
    else:
        subj_scores_div = student_finals
    division_points, division = compute_division_from_finals(sid, subj_scores_div)

    return jsonify({"ok":True,"student":student,"term":term,"results":results,"ca_count":ca_count,
                    "average":avg,"grade":get_grade(sid,avg),
                    "class_position":c_pos,"class_total":c_total,
                    "stream_position":s_pos,"stream_total":s_total,"assess":assess,
                    "division":division,"division_points":division_points})

# ── PLATFORM ANNOUNCEMENTS (school-side) ──────────────────────
@app.route("/api/platform_announcements", methods=["GET"])
@require_auth
def api_platform_announcements():
    sid=g.school_id
    con=get_db(); cur=con.cursor()
    cur.execute("""SELECT pa.id,pa.title,pa.body,CAST(pa.posted_at AS TEXT) as posted_at,
                          CASE WHEN par.school_id IS NOT NULL THEN 1 ELSE 0 END AS is_read
                   FROM platform_announcements pa
                   LEFT JOIN platform_announcement_reads par ON par.announcement_id=pa.id AND par.school_id=%s
                   WHERE pa.target='all' OR pa.target=%s
                   ORDER BY pa.posted_at DESC""",(sid,str(sid)))
    rows=to_dicts(cur.fetchall(),cur); cur.close(); con.close(); return jsonify(rows)

@app.route("/api/platform_announcements/<int:aid>/read", methods=["POST"])
@require_auth
def api_platform_announcement_read(aid):
    sid=g.school_id
    con=get_db(); cur=con.cursor()
    cur.execute("INSERT INTO platform_announcement_reads(announcement_id,school_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",(aid,sid))
    con.commit(); cur.close(); con.close(); return jsonify({"ok":True})

@app.route("/api/subscription/status", methods=["GET"])
@require_auth
@require_role("admin")
def api_subscription_status():
    sid = g.school_id
    _expire_stale_payment_requests(sid)
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT subscription_exempt, subscription_status, subscription_plan, subscription_expires_at FROM schools WHERE id=%s", (sid,))
    row = cur.fetchone()
    if not row:
        cur.close(); con.close(); return jsonify({"ok": False}), 404
    exempt, status, plan, expires_at = row

    cur.execute("""SELECT id, plan, claimed_amount, transaction_id, phone_used,
                          CAST(payment_date AS TEXT), note, status, CAST(submitted_at AS TEXT), decision_note
                   FROM payment_requests WHERE school_id=%s ORDER BY id DESC LIMIT 1""", (sid,))
    req_row = cur.fetchone(); cur.close(); con.close()

    pending_request = None
    last_decision = None
    if req_row:
        (rid, rplan, ramount, rtxn, rphone, rdate, rnote, rstatus, rsubmitted, rdecision_note) = req_row
        if rstatus == "pending":
            pending_request = {"id": rid, "plan": rplan, "claimed_amount": ramount,
                                "transaction_id": rtxn, "phone_used": rphone,
                                "payment_date": rdate, "note": rnote, "submitted_at": rsubmitted}
        elif rstatus == "expired":
            last_decision = {"status": "expired",
                              "note": "This payment request expired because it was not verified within 48 hours. "
                                      "If you already paid, contact support. Otherwise, submit a new payment request."}
        elif rstatus == "rejected":
            last_decision = {"status": "rejected", "note": rdecision_note}

    return jsonify({"ok": True, "active": is_subscribed(sid), "exempt": bool(exempt),
                     "plan": plan, "status": status,
                     "expires_at": expires_at.isoformat() if expires_at else None,
                     "plans": SUBSCRIPTION_PLANS,
                     "payment_info": _platform_payment_config(),
                     "pending_request": pending_request,
                     "last_decision": last_decision})
@app.route("/api/subscription/select_free", methods=["POST"])
@require_auth
@require_role("admin")
def api_select_free_plan():
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("""UPDATE schools SET subscription_status='active', subscription_plan='free',
                   subscription_expires_at=%s WHERE id=%s""",
                (datetime.utcnow() + timedelta(days=SUBSCRIPTION_PLANS["free"]["days"]), sid))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok": True})

@app.route("/api/subscription/request", methods=["POST"])
@require_auth
@require_role("admin")
def api_submit_payment_request():
    sid = g.school_id
    _expire_stale_payment_requests(sid)
    d = request.json or {}
    plan = d.get("plan", "")
    if plan not in SUBSCRIPTION_PLANS or plan == "free":
        return jsonify({"ok": False, "error": "Invalid plan"}), 400
    txn_id  = (d.get("transaction_id") or "").strip()
    phone   = (d.get("phone_used") or "").strip()
    pay_date= (d.get("payment_date") or "").strip()
    note    = (d.get("note") or "").strip()
    try:
        amount = float(d.get("claimed_amount"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Amount paid is required"}), 400
    if not txn_id:   return jsonify({"ok": False, "error": "Transaction ID is required"}), 400
    if not phone:    return jsonify({"ok": False, "error": "Phone number used is required"}), 400
    if amount <= 0:  return jsonify({"ok": False, "error": "Amount paid must be greater than zero"}), 400
    if not pay_date: return jsonify({"ok": False, "error": "Payment date is required"}), 400

    con = get_db(); cur = con.cursor()

    # Abuse guard: counts every submission attempt (any status) in the last
    # 24h, so rapid cancel-and-resubmit cycling still counts against the cap.
    cur.execute("SELECT COUNT(*) FROM payment_requests WHERE school_id=%s AND submitted_at > NOW() - INTERVAL '24 hours'", (sid,))
    if cur.fetchone()[0] >= 5:
        cur.close(); con.close()
        return jsonify({"ok": False, "error": "Too many verification requests. Please contact support if you continue experiencing problems."}), 429

    cur.execute("SELECT id FROM payment_requests WHERE school_id=%s AND status='pending'", (sid,))
    if cur.fetchone():
        cur.close(); con.close()
        return jsonify({"ok": False, "error": "You already have a payment request pending verification. Cancel it below before submitting a new one."}), 409
    try:
        cur.execute("""INSERT INTO payment_requests(school_id,plan,claimed_amount,transaction_id,phone_used,payment_date,note,status,submitted_by)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,'pending',%s) RETURNING id""",
                    (sid, plan, amount, txn_id, phone, pay_date, note, d.get("username", "")))
        new_id = cur.fetchone()[0]
        cur.execute("UPDATE schools SET subscription_status='pending' WHERE id=%s", (sid,))
        con.commit()
    except psycopg2.errors.UniqueViolation:
        con.rollback(); cur.close(); con.close()
        return jsonify({"ok": False, "error": "This transaction ID has already been submitted. If you believe this is an error, contact support."}), 409
    cur.close(); con.close()
    return jsonify({"ok": True, "id": new_id})

@app.route("/api/subscription/request/cancel", methods=["POST"])
@require_auth
@require_role("admin")
def api_cancel_payment_request():
    sid = g.school_id
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT id FROM payment_requests WHERE school_id=%s AND status='pending' ORDER BY id DESC LIMIT 1", (sid,))
    row = cur.fetchone()
    if not row:
        cur.close(); con.close(); return jsonify({"ok": False, "error": "No pending request to cancel"}), 404
    cur.execute("UPDATE payment_requests SET status='cancelled', decided_at=NOW() WHERE id=%s", (row[0],))
    cur.execute("UPDATE schools SET subscription_status='inactive' WHERE id=%s AND subscription_status='pending'", (sid,))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok": True})

# ── PDF ────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, A3, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

def _get_logo_element(school_id, max_h=2*cm):
    logo_data = get_config_val(school_id, "logo_data", "")
    if logo_data:
        try:
            raw = base64.b64decode(logo_data)
            img = Image(io.BytesIO(raw)); img.drawHeight=max_h; img.drawWidth=max_h; return img
        except: pass
    # Fallback for any older logo still sitting on disk (e.g. one committed to Git)
    path = get_config_val(school_id,"logo_path","")
    if path and not path.startswith("api/logo/"):
        full = os.path.join(BASE_DIR, path)
        if os.path.exists(full):
            try:
                img=Image(full); img.drawHeight=max_h; img.drawWidth=max_h; return img
            except: pass
    return None

def _esc(s):
    return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

def _school_header_story(school_id, styles, title_text, subtitle_text=""):
    story=[]; H_BG=colors.HexColor("#1A6FA8")
    t_s=ParagraphStyle("T",parent=styles["Title"],fontSize=16,textColor=H_BG,spaceAfter=2,alignment=1)
    s_s=ParagraphStyle("S",parent=styles["Normal"],fontSize=9,alignment=1,spaceAfter=4)
    motto=get_config_val(school_id,"motto",""); sname=get_school_name(school_id)
    logo=_get_logo_element(school_id,1.8*cm)
    if logo:
        name_para=Paragraph(f"<b>{sname}</b>",t_s)
        sub_para=Paragraph(motto,s_s) if motto else None
        inner=[[logo,[name_para]+([sub_para] if sub_para else [])]]
        tbl=Table(inner,colWidths=[2.2*cm,None])
        tbl.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),6)]))
        story.append(tbl)
    else:
        story.append(Paragraph(sname,t_s))
        if motto: story.append(Paragraph(f'<i>"{motto}"</i>',s_s))
    if title_text: story.append(Paragraph(title_text,s_s))
    if subtitle_text: story.append(Paragraph(subtitle_text,s_s))
    story.append(Spacer(1,0.3*cm)); return story

def _blue_sheet_pdf(school_id,filename,subtitle,students,subjects,get_score_fn,term=None,class_label=""):
    subj_map=get_subject_map(school_id)
    H_BG=colors.HexColor("#1A6FA8"); S_BG=colors.HexColor("#5BA4CF")
    ODD=colors.HexColor("#E8F4FC"); EVEN=colors.white
    RED=colors.HexColor("#C0392B"); WHITE=colors.white
    page_size = landscape(A3) if len(subjects) > 10 else landscape(A4)
    doc=SimpleDocTemplate(filename,pagesize=page_size,rightMargin=1.2*cm,leftMargin=1.2*cm,topMargin=1.2*cm,bottomMargin=1.2*cm)
    styles=getSampleStyleSheet(); story=[]
    tl=term["label"] if term else ""
    title_text = f"{subtitle} — {class_label}" if class_label else subtitle
    story+=_school_header_story(school_id,styles,title_text,tl)
    results=[]
    for s in students:
        row={"name":s["name"],"scores":{},"total":0,"count":0}
        for subj in subjects:
            sc=get_score_fn(s["id"],subj); row["scores"][subj]=sc
            if sc is not None: row["total"]+=sc; row["count"]+=1
        row["average"]=row["total"]/row["count"] if row["count"] else 0
        row["grade"]=get_grade(school_id,row["average"]); results.append(row)
    _assign_positions(results,"average")
    name_style = ParagraphStyle("NameCell", parent=styles["Normal"], fontSize=7.5, leading=8.5)
    hdr=["#","Student"]+[subj_map.get(s,s[:4].upper()) for s in subjects]+["Total","Avg","Pos","Grd"]
    tdata=[hdr]; fail_cells=[]
    for ri,r in enumerate(results,1):
        row=[str(ri),Paragraph(_esc(r["name"]),name_style)]
        for ci,subj in enumerate(subjects):
            sc=r["scores"][subj]
            if sc is not None:
                if sc<50: fail_cells.append((ri,ci+2))
                row.append(f"{sc:.1f}")
            else: row.append("-")
        row+=[f"{r['total']:.1f}" if r["count"] else "-",
              f"{r['average']:.1f}" if r["count"] else "-",
              str(r["position"]),r["grade"] if r["count"] else "-"]
        tdata.append(row)
    sc_w=1.2*cm
    cw=[0.8*cm,6.2*cm]+[sc_w]*len(subjects)+[1.5*cm,1.3*cm,0.9*cm,1.0*cm]
    tbl=Table(tdata,colWidths=cw,repeatRows=1)
    ts=[("BACKGROUND",(0,0),(-1,0),H_BG),("TEXTCOLOR",(0,0),(-1,0),WHITE),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),6.5),
        ("ALIGN",(0,0),(-1,0),"CENTER"),("FONTNAME",(0,1),(-1,-1),"Helvetica"),
        ("FONTSIZE",(0,1),(-1,-1),7),("ALIGN",(0,1),(-1,-1),"CENTER"),("ALIGN",(1,1),(1,-1),"LEFT"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[ODD,EVEN]),("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#A0C4E0")),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("BACKGROUND",(-4,0),(-1,0),S_BG),("FONTNAME",(-4,1),(-1,-1),"Helvetica-Bold")]
    for (ri,ci) in fail_cells: ts.append(("TEXTCOLOR",(ci,ri),(ci,ri),RED))
    tbl.setStyle(TableStyle(ts)); story.append(tbl)
    story.append(Spacer(1,0.3*cm))
    ft=ParagraphStyle("F",parent=styles["Normal"],fontSize=6.5,textColor=colors.grey,alignment=2)
    story.append(Paragraph(f"Generated | {len(students)} students | {tl}",ft))
    doc.build(story)

@app.route("/api/pdf/report/<int:sid>", methods=["GET"])
@subscription_required
@require_auth
def pdf_report(sid):
    school_id=g.school_id; subjects=get_subjects(school_id); subj_map=get_subject_map(school_id)
    term_id=request.args.get("term_id")
    con=get_db(); cur=con.cursor()
    cur.execute("""SELECT s.id,s.name,s.class_id,s.stream_id,c.class_name,st.stream_name
                   FROM students s JOIN classes c ON s.class_id=c.id LEFT JOIN streams st ON s.stream_id=st.id
                   WHERE s.id=%s AND s.school_id=%s""",(sid,school_id))
    row=cur.fetchone(); student=to_dict(row,cur) if row else None; cur.close(); con.close()
    if not student: return jsonify({"error":"Not found"}),404
    term=get_term_by_id(school_id,int(term_id)) if term_id else get_active_term(school_id)
    if not term: return jsonify({"error":"No term"}),400
    tid=term["id"]; ca_count=term["ca_count"]; ca_w=term["ca_weight"]; ex_w=term["exam_weight"]
    class_id=student["class_id"]; stream_id=student["stream_id"]

    class_rows, class_rank_map, stream_rank_map, scores_bulk = get_class_report_data(
        school_id, tid, class_id, stream_id, subjects, ca_w, ex_w)
    subject_rank_maps = {subj: get_subject_rank_map(class_rows, subj) for subj in subjects}
    c_entry=class_rank_map.get(sid)
    c_pos=c_entry["position"] if c_entry else "-"
    c_total=len(class_rows)
    s_pos=s_total=None
    if stream_id and stream_rank_map is not None:
        s_entry=stream_rank_map.get(sid)
        s_pos=s_entry["position"] if s_entry else "-"
        s_total=len(stream_rank_map)
    student_finals = c_entry["finals"] if c_entry else compute_student_finals(scores_bulk, sid, subjects, ca_w, ex_w)
    avg = c_entry["average"] if c_entry else compute_average_from_finals(student_finals)
    student_scores = scores_bulk.get(sid, {})

    safe=student["name"].replace(" ","_")
    fname=os.path.join(tempfile.gettempdir(),f"RC_{school_id}_{safe}_{tid}.pdf")
    # Many CA columns get cramped in portrait — switch to landscape automatically.
    page_size = landscape(A4) if ca_count > 4 else A4
    doc=SimpleDocTemplate(fname,pagesize=page_size,rightMargin=1.5*cm,leftMargin=1.5*cm,topMargin=1.5*cm,bottomMargin=1.5*cm)
    styles=getSampleStyleSheet(); story=[]
    H_BG=colors.HexColor("#1A6FA8"); ODD=colors.HexColor("#E8F4FC"); WHITE=colors.white; RED=colors.HexColor("#C0392B")
    story+=_school_header_story(school_id,styles,"STUDENT REPORT CARD")
    stream_label=f"{student['class_name']} {student['stream_name']}" if student.get("stream_name") else student["class_name"]
    info=[[" Name:",student["name"],"Class:",stream_label],
          ["Term:",term["label"],"Weights:",f"CA {ca_w}% | Exam {ex_w}%"],
          ["Class Position:",f"{c_pos}/{c_total}","Grade:",get_grade(school_id,avg)]]
    if s_pos is not None: info.append(["Stream Position:",f"{s_pos}/{s_total}","",""])
    it=Table(info,colWidths=[3*cm,6*cm,3*cm,5*cm])
    it.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),"Helvetica"),("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
        ("FONTNAME",(2,0),(2,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]))
    story+=[it,Spacer(1,0.4*cm)]
    hdr=["Subject"]+[f"CA{i}" for i in range(1,ca_count+1)]+["Exam","Final","Pos","Grd","Remark","Sign"]
    tdata=[hdr]; tot,cnt=0,0; fail_rows=[]
    for subject in subjects:
        entry=student_scores.get(subject,{})
        ca_map=entry.get("ca",{})
        row_data=[subject.title()]+[f"{ca_map.get(f'CA{i}'):.1f}" if ca_map.get(f"CA{i}") is not None else "-" for i in range(1,ca_count+1)]
        exam_v=entry.get("exam")
        row_data.append(f"{exam_v:.1f}" if exam_v is not None else "-")
        final_v=student_finals.get(subject)
        if final_v is not None: tot+=final_v; cnt+=1
        row_data.append(f"{final_v:.1f}" if final_v is not None else "-")
        row_data.append(str(subject_rank_maps[subject].get(sid,"-")) if final_v is not None else "-")
        row_data.append(get_grade(school_id,final_v) if final_v is not None else "-")
        row_data+=["",""]
        if final_v is not None and final_v<50: fail_rows.append(len(tdata))
        tdata.append(row_data)
    ca_cw=1.1*cm; cw=[4.0*cm]+[ca_cw]*ca_count+[1.4*cm,1.4*cm,1.0*cm,1.1*cm,2.4*cm,1.6*cm]
    mt=Table(tdata,colWidths=cw,repeatRows=1)
    mts=[("BACKGROUND",(0,0),(-1,0),H_BG),("TEXTCOLOR",(0,0),(-1,0),WHITE),
         ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7.5),
         ("ALIGN",(0,0),(-1,-1),"CENTER"),("ALIGN",(0,1),(0,-1),"LEFT"),
         ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#A0C4E0")),("ROWBACKGROUNDS",(0,1),(-1,-1),[ODD,WHITE]),
         ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]
    for ri in fail_rows: mts.append(("TEXTCOLOR",(0,ri),(-1,ri),RED))
    mt.setStyle(TableStyle(mts)); story+=[mt,Spacer(1,0.4*cm)]
    comp_avg=tot/cnt if cnt else 0
    summary_data=[["AVERAGE",f"{comp_avg:.2f}","GRADE",get_grade(school_id,comp_avg),"CLASS POS",f"{c_pos}/{c_total}"]]
    summary_cols=[3*cm,3*cm,2*cm,2*cm,3*cm,4*cm]
    if s_pos is not None: summary_data[0]+=["STREAM POS",f"{s_pos}/{s_total}"]; summary_cols+=[3*cm,3*cm]
    sm=Table(summary_data,colWidths=summary_cols)
    sm.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),H_BG),("TEXTCOLOR",(0,0),(-1,-1),WHITE),
        ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    story+=[sm,Spacer(1,0.4*cm)]
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT * FROM remarks WHERE school_id=%s AND student_id=%s AND term_id=%s",(school_id,sid,tid))
    rmk_row=cur.fetchone(); rmk=to_dict(rmk_row,cur) if rmk_row else None; cur.close(); con.close()
    rm_data=[["Class Teacher Remark:",rmk["class_teacher_remark"] if rmk else "________________________"],
             ["Head of School Remark:",rmk["head_remark"] if rmk else "________________________"]]
    rmt=Table(rm_data,colWidths=[5*cm,12*cm])
    rmt.setStyle(TableStyle([("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4),("LINEBELOW",(1,0),(1,-1),0.5,colors.grey)]))
    story+=[rmt,Spacer(1,0.4*cm)]
    sig=Table([["Class Teacher Sign: _______________","Head Sign: _______________","Date: _______________"]],colWidths=[6*cm,6*cm,5*cm])
    sig.setStyle(TableStyle([("FONTSIZE",(0,0),(-1,-1),7.5),("FONTNAME",(0,0),(-1,-1),"Helvetica")]))
    story.append(sig); doc.build(story)
    return send_file(fname,as_attachment=True,
                     download_name=f"RC_{student['name'].replace(' ','_')}_{term['label'].replace(' ','_')}.pdf",
                     mimetype="application/pdf")

@app.route("/api/pdf/ca_sheet", methods=["GET"])
@subscription_required
@require_auth
def pdf_ca_sheet():
    school_id=g.school_id; subjects=get_subjects(school_id)
    class_id=request.args.get("class_id"); stream_id=request.args.get("stream_id") or None
    ca_name=request.args.get("ca_name","CA1"); term_id=request.args.get("term_id")
    term=get_term_by_id(school_id,int(term_id)) if term_id else get_active_term(school_id)
    if not term: return jsonify({"error":"No term"}),400
    tid=term["id"]; class_id=int(class_id) if class_id else None
    if stream_id: stream_id=int(stream_id)
    studs=get_students_in_scope(school_id,class_id,stream_id)
    if not studs: return jsonify({"error":"No students"}),404
    fname=os.path.join(tempfile.gettempdir(),f"CA_{school_id}_{class_id}_{ca_name.replace(':','_')}_{tid}.pdf")
    scores_bulk=get_term_scores_bulk(school_id,tid,[s["id"] for s in studs])
    def get_score(stid,subj):
        entry=scores_bulk.get(stid,{}).get(subj)
        if not entry: return None
        if ca_name=="exam": return entry["exam"]
        if ca_name.startswith("test:"):
            tid_=int(ca_name.split(":",1)[1])
            return (entry.get("tests") or {}).get(tid_)
        return entry["ca"].get(ca_name)
    subtitle_label = ca_name.upper()
    if ca_name.startswith("test:"):
        con=get_db(); cur=con.cursor()
        cur.execute("SELECT label FROM term_tests WHERE id=%s AND school_id=%s",(int(ca_name.split(":",1)[1]),school_id))
        r=cur.fetchone(); cur.close(); con.close()
        subtitle_label = (r[0].upper() if r else "TEST")
    class_label = studs[0]["class_name"] + (f" {studs[0]['stream_name']}" if stream_id and studs[0].get("stream_name") else "")
    _blue_sheet_pdf(school_id,fname,f"{subtitle_label} SCORE SHEET",studs,subjects,get_score,term,class_label=class_label)
    return send_file(fname,as_attachment=True,download_name=os.path.basename(fname),mimetype="application/pdf")

@app.route("/api/pdf/grade_sheet", methods=["GET"])
@subscription_required
@require_auth
def pdf_grade_sheet():
    school_id=g.school_id; subjects=get_subjects(school_id)
    mode=request.args.get("mode","ca")
    class_id=request.args.get("class_id"); stream_id=request.args.get("stream_id") or None
    ca_name=request.args.get("ca_name","CA1"); term_id=request.args.get("term_id")
    grading_system=request.args.get("grading_system") or None
    division_source=request.args.get("division_source") or None
    noncredit_param=request.args.get("noncredit","")
    noncredit_override=[x.strip().lower() for x in noncredit_param.split(",") if x.strip()] if noncredit_param else None

    term=get_term_by_id(school_id,int(term_id)) if term_id else get_active_term(school_id)
    if not term: return jsonify({"error":"No term"}),400
    tid=term["id"]; class_id=int(class_id) if class_id else None
    if stream_id: stream_id=int(stream_id)
    studs=get_students_in_scope(school_id,class_id,stream_id)
    if not studs: return jsonify({"error":"No students"}),404

    scores_bulk=get_term_scores_bulk(school_id,tid,[s["id"] for s in studs])
    ca_w=term["ca_weight"]; ex_w=term["exam_weight"]
    def get_score(stid,subj):
        entry=scores_bulk.get(stid,{}).get(subj)
        if not entry: return None
        if mode=="ca":
            if ca_name.startswith("test:"):
                tid_=int(ca_name.split(":",1)[1])
                return (entry.get("tests") or {}).get(tid_)
            return entry["ca"].get(ca_name)
        if mode=="exam": return entry["exam"]
        if mode=="terminal": return _final_from_entry(entry,ca_w,ex_w)
        return None

    settings = get_school_grading_settings(school_id)
    level = grading_system or settings["grading_system"]
    div_source = division_source or settings["division_source"]
    rules = get_necta_grades(level) if div_source=="necta" else get_grade_rules(school_id)

    fname=os.path.join(tempfile.gettempdir(),f"Grade_{school_id}_{class_id}_{mode}_{ca_name.replace(':','_')}_{tid}.pdf")
    class_label = studs[0]["class_name"] + (f" {studs[0]['stream_name']}" if stream_id and studs[0].get("stream_name") else "")
    if mode=="ca" and ca_name.startswith("test:"):
        con=get_db(); cur=con.cursor()
        cur.execute("SELECT label FROM term_tests WHERE id=%s AND school_id=%s",(int(ca_name.split(":",1)[1]),school_id))
        r=cur.fetchone(); cur.close(); con.close()
        title = f"{(r[0].upper() if r else 'TEST')} GRADE SHEET"
    else:
        title = {"ca": f"{ca_name.upper()} GRADE SHEET", "exam":"EXAM GRADE SHEET", "terminal":"TERMINAL GRADE SHEET"}.get(mode,"GRADE SHEET")

    subj_map=get_subject_map(school_id)
    H_BG=colors.HexColor("#1A6FA8"); ODD=colors.HexColor("#E8F4FC"); WHITE=colors.white
    page_size = landscape(A3) if len(subjects) > 10 else landscape(A4)
    doc=SimpleDocTemplate(fname,pagesize=page_size,rightMargin=1.2*cm,leftMargin=1.2*cm,topMargin=1.2*cm,bottomMargin=1.2*cm)
    styles=getSampleStyleSheet(); story=[]
    story+=_school_header_story(school_id,styles,f"{title} — {class_label}" if class_label else title, term["label"])

    name_style = ParagraphStyle("NameCell", parent=styles["Normal"], fontSize=7.5, leading=8.5)
    hdr=["#","Student"]+[subj_map.get(s,s[:4].upper()) for s in subjects]+["Points","Division"]
    tdata=[hdr]
    for ri,s in enumerate(studs,1):
        subj_scores={}
        row=[str(ri),Paragraph(_esc(s["name"]),name_style)]
        for subj in subjects:
            sc=get_score(s["id"],subj)
            subj_scores[subj]=sc
            grade,_ = grade_and_points_for_score(rules, sc)
            row.append(grade)
        points, division = compute_division_from_finals(school_id, subj_scores, grading_system, division_source, noncredit_override)
        row += [str(points) if points is not None else "-", division or "-"]
        tdata.append(row)

    sc_w=1.2*cm
    cw=[0.8*cm,6.2*cm]+[sc_w]*len(subjects)+[1.5*cm,1.5*cm]
    tbl=Table(tdata,colWidths=cw,repeatRows=1)
    ts=[("BACKGROUND",(0,0),(-1,0),H_BG),("TEXTCOLOR",(0,0),(-1,0),WHITE),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),7),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),("ALIGN",(1,1),(1,-1),"LEFT"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[ODD,WHITE]),("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#A0C4E0")),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]
    tbl.setStyle(TableStyle(ts)); story.append(tbl)
    story.append(Spacer(1,0.3*cm))
    ft=ParagraphStyle("F",parent=styles["Normal"],fontSize=6.5,textColor=colors.grey,alignment=2)
    story.append(Paragraph(f"Generated | {len(studs)} students | {term['label']}",ft))
    doc.build(story)
    return send_file(fname,as_attachment=True,download_name=os.path.basename(fname),mimetype="application/pdf")

@app.route("/api/pdf/terminal_sheet", methods=["GET"])
@subscription_required
@require_auth
def pdf_terminal_sheet():
    school_id=g.school_id; subjects=get_subjects(school_id)
    class_id=request.args.get("class_id"); stream_id=request.args.get("stream_id") or None; term_id=request.args.get("term_id")
    term=get_term_by_id(school_id,int(term_id)) if term_id else get_active_term(school_id)
    if not term: return jsonify({"error":"No term"}),400
    tid=term["id"]; class_id=int(class_id) if class_id else None
    if stream_id: stream_id=int(stream_id)
    studs=get_students_in_scope(school_id,class_id,stream_id)
    if not studs: return jsonify({"error":"No students"}),404
    fname=os.path.join(tempfile.gettempdir(),f"Terminal_{school_id}_{class_id}_{tid}.pdf")
    scores_bulk=get_term_scores_bulk(school_id,tid,[s["id"] for s in studs])
    ca_w=term["ca_weight"]; ex_w=term["exam_weight"]
    def get_score(stid,subj):
        f=_final_from_entry(scores_bulk.get(stid,{}).get(subj),ca_w,ex_w)
        return round(f,1) if f is not None else None
    class_label = studs[0]["class_name"] + (f" {studs[0]['stream_name']}" if stream_id and studs[0].get("stream_name") else "")
    _blue_sheet_pdf(school_id,fname,f"TERMINAL SCORE SHEET (CA {term['ca_weight']}% + Exam {term['exam_weight']}%)",studs,subjects,get_score,term)
    return send_file(fname,as_attachment=True,download_name=os.path.basename(fname),mimetype="application/pdf")


# ══════════════════════════════════════════════════════════════
# SUPERADMIN ROUTES
# ══════════════════════════════════════════════════════════════
_SA_SESSIONS = {}

def _sa_token():
    auth=request.headers.get("Authorization","")
    if auth.startswith("Bearer "): return auth[7:].strip()
    return None

def _require_superadmin():
    token=_sa_token()
    if not token or token not in _SA_SESSIONS:
        return None,(jsonify({"ok":False,"error":"Superadmin authentication required"}),401)
    return _SA_SESSIONS[token],None

@app.route("/api/superadmin/login", methods=["POST"])
def api_superadmin_login():
    d=request.json; u=d.get("username","").strip(); p=d.get("password","")
    if not u or not p: return jsonify({"ok":False,"error":"Username and password required"}),400
    con=get_db(); cur=con.cursor()
    cur.execute("SELECT password FROM superadmins WHERE username=%s",(u,))
    row=cur.fetchone(); cur.close(); con.close()
    if not row or not verify_password(p,row[0]): return jsonify({"ok":False,"error":"Invalid credentials"}),401
    token=secrets.token_hex(32); _SA_SESSIONS[token]=u
    return jsonify({"ok":True,"token":token,"username":u})

@app.route("/api/superadmin/logout", methods=["POST"])
def api_superadmin_logout():
    token=_sa_token()
    if token and token in _SA_SESSIONS: del _SA_SESSIONS[token]
    return jsonify({"ok":True})

@app.route("/api/superadmin/schools", methods=["GET"])
def api_superadmin_schools():
    sa,err=_require_superadmin()
    if err: return err
    try:
        con=get_db(); cur=con.cursor()
        cur.execute("SELECT id,school_name,CAST(registered_at AS TEXT) FROM schools ORDER BY id")
        schools_raw=to_dicts(cur.fetchall(),cur); result=[]
        for s in schools_raw:
            sid=s["id"]
            cur.execute("SELECT COUNT(*) FROM students WHERE school_id=%s",(sid,)); sc=cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM users WHERE school_id=%s AND role='teacher'",(sid,)); tc=cur.fetchone()[0]
            cur.execute("SELECT label FROM terms WHERE school_id=%s AND status='open' ORDER BY id DESC LIMIT 1",(sid,))
            r=cur.fetchone(); at=r[0] if r else "—"
            cur.execute("SELECT subscription_exempt, subscription_status, subscription_plan, subscription_expires_at FROM schools WHERE id=%s",(sid,))
            sub_row = cur.fetchone()
            exempt, sub_status, sub_plan, sub_expires = sub_row if sub_row else (0,"inactive","",None)
            payment_status = "demo" if exempt else (sub_status or "inactive")
            result.append({"id":sid,"school_name":s["school_name"],"registered_at":s.get("cast") or "—",
                           "student_count":sc,"teacher_count":tc,"active_term":at,
                           "payment_status":payment_status,
                           "subscription_plan": sub_plan or "—",
                           "subscription_expires_at": sub_expires.isoformat() if sub_expires else None})
        cur.close(); con.close()
        return jsonify({"ok":True,"schools":result})
    except Exception as e: return jsonify({"ok":False,"error":str(e)}),500

@app.route("/api/superadmin/announce", methods=["GET"])
def api_superadmin_announce_list():
    sa,err=_require_superadmin()
    if err: return err
    try:
        con=get_db(); cur=con.cursor()
        cur.execute("SELECT id,title,body,target,CAST(posted_at AS TEXT) as posted_at FROM platform_announcements ORDER BY posted_at DESC")
        rows=to_dicts(cur.fetchall(),cur); cur.close(); con.close()
        return jsonify({"ok":True,"announcements":rows})
    except Exception as e: return jsonify({"ok":False,"error":str(e)}),500

@app.route("/api/superadmin/announce", methods=["POST"])
def api_superadmin_announce():
    sa,err=_require_superadmin()
    if err: return err
    d=request.json; title=d.get("title","").strip(); body=d.get("body","").strip(); target=d.get("target","all").strip()
    if not title or not body: return jsonify({"ok":False,"error":"Title and body required"}),400
    try:
        con=get_db(); cur=con.cursor()
        cur.execute("INSERT INTO platform_announcements(title,body,target) VALUES(%s,%s,%s) RETURNING id",(title,body,target))
        new_id=cur.fetchone()[0]; con.commit(); cur.close(); con.close()
        return jsonify({"ok":True,"id":new_id})
    except Exception as e: return jsonify({"ok":False,"error":str(e)}),500

@app.route("/api/superadmin/announce/<int:aid>", methods=["DELETE"])
def api_superadmin_announce_delete(aid):
    sa,err=_require_superadmin()
    if err: return err
    try:
        con=get_db(); cur=con.cursor()
        cur.execute("DELETE FROM platform_announcements WHERE id=%s",(aid,))
        con.commit(); cur.close(); con.close()
        return jsonify({"ok":True})
    except Exception as e: return jsonify({"ok":False,"error":str(e)}),500

@app.route("/api/reset_db", methods=["POST"])
def api_reset_db():
    secret=request.json.get("secret","")
    if secret!=os.environ.get("ADMIN_SETUP_SECRET",""): return jsonify({"ok":False,"error":"Invalid secret"}),403
    con=get_db(); cur=con.cursor()
    for t in ["remarks","exam_scores","ca_scores","subject_assignments","students","streams","classes",
              "terms","grade_config","school_subjects","school_config","announcements","announcement_reads",
              "results_published","users","schools"]:
        cur.execute(f"DROP TABLE IF EXISTS {t} CASCADE")
    con.commit(); cur.close(); con.close(); init_db()
    return jsonify({"ok":True,"message":"Database reset."})

@app.route("/api/superadmin/payment_requests", methods=["GET"])
def api_sa_payment_requests():
    sa, err = _require_superadmin()
    if err: return err
    _expire_stale_payment_requests()
    status = request.args.get("status", "pending")
    con = get_db(); cur = con.cursor()
    q = """SELECT pr.id, pr.school_id, s.school_name, pr.plan, pr.claimed_amount, pr.transaction_id,
                  pr.phone_used, CAST(pr.payment_date AS TEXT), pr.note, pr.status,
                  pr.submitted_by, CAST(pr.submitted_at AS TEXT),
                  pr.decided_by, CAST(pr.decided_at AS TEXT), pr.decision_note
           FROM payment_requests pr JOIN schools s ON s.id = pr.school_id"""
    params = ()
    if status != "all":
        q += " WHERE pr.status=%s"; params = (status,)
    q += " ORDER BY pr.submitted_at ASC"
    cur.execute(q, params)
    rows = to_dicts(cur.fetchall(), cur); cur.close(); con.close()
    return jsonify({"ok": True, "requests": rows})

@app.route("/api/superadmin/payment_requests/<int:rid>/approve", methods=["POST"])
def api_sa_approve_payment(rid):
    sa, err = _require_superadmin()
    if err: return err
    note = (request.json or {}).get("note", "")
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT school_id, plan, status FROM payment_requests WHERE id=%s", (rid,))
    row = cur.fetchone()
    if not row: cur.close(); con.close(); return jsonify({"ok": False, "error": "Not found"}), 404
    school_id, plan, status = row
    if status not in ("pending", "expired"):
        cur.close(); con.close(); return jsonify({"ok": False, "error": "This request was already decided"}), 409
    if plan not in SUBSCRIPTION_PLANS:
        cur.close(); con.close(); return jsonify({"ok": False, "error": "Unknown plan on this request"}), 400
    days = SUBSCRIPTION_PLANS[plan]["days"]
    cur.execute("SELECT subscription_expires_at FROM schools WHERE id=%s", (school_id,))
    exp_row = cur.fetchone()
    base = exp_row[0] if exp_row and exp_row[0] and exp_row[0] > datetime.utcnow() else datetime.utcnow()
    cur.execute("""UPDATE schools SET subscription_status='active', subscription_plan=%s,
                   subscription_expires_at=%s WHERE id=%s""",
                (plan, base + timedelta(days=days), school_id))
    cur.execute("""UPDATE payment_requests SET status='approved', decided_by=%s, decided_at=NOW(), decision_note=%s
                   WHERE id=%s""", (sa, note, rid))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok": True})

@app.route("/api/superadmin/payment_requests/<int:rid>/reject", methods=["POST"])
def api_sa_reject_payment(rid):
    sa, err = _require_superadmin()
    if err: return err
    note = (request.json or {}).get("note", "").strip()
    if not note: return jsonify({"ok": False, "error": "A rejection reason is required"}), 400
    con = get_db(); cur = con.cursor()
    cur.execute("SELECT school_id, status FROM payment_requests WHERE id=%s", (rid,))
    row = cur.fetchone()
    if not row: cur.close(); con.close(); return jsonify({"ok": False, "error": "Not found"}), 404
    school_id, status = row
    if status not in ("pending", "expired"):
        cur.close(); con.close(); return jsonify({"ok": False, "error": "This request was already decided"}), 409
    cur.execute("UPDATE payment_requests SET status='rejected', decided_by=%s, decided_at=NOW(), decision_note=%s WHERE id=%s",
                (sa, note, rid))
    cur.execute("UPDATE schools SET subscription_status='inactive' WHERE id=%s AND subscription_status='pending'", (school_id,))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok": True})

@app.route("/api/superadmin/payment_config", methods=["GET"])
def api_sa_get_payment_config():
    sa, err = _require_superadmin()
    if err: return err
    return jsonify({"ok": True, **_platform_payment_config()})

@app.route("/api/superadmin/payment_config", methods=["POST"])
def api_sa_set_payment_config():
    sa, err = _require_superadmin()
    if err: return err
    d = request.json or {}
    business_name = (d.get("business_name") or "").strip()
    payment_number = (d.get("payment_number") or "").strip()
    networks = d.get("networks") or []
    if not business_name or not payment_number:
        return jsonify({"ok": False, "error": "Business name and payment number required"}), 400
    con = get_db(); cur = con.cursor()
    cur.execute("""UPDATE platform_payment_config SET business_name=%s, payment_number=%s, networks=%s, updated_at=NOW() WHERE id=1""",
                (business_name, payment_number, ",".join(networks)))
    con.commit(); cur.close(); con.close()
    return jsonify({"ok": True})

# ── STATIC ─────────────────────────────────────────────────────
_STATIC_FILES=["shared.css","shared.js","page-dashboard.js","page-students.js",
               "page-teachers.js","page-reports.js","page-parent.js","page-config.js",
               "page-analytics.js"]

@app.route("/")
def index(): return send_from_directory(BASE_DIR,"index.html")

@app.route("/setup")
def setup_page(): return send_from_directory(BASE_DIR,"setup.html")

@app.route("/register")
def register_page(): return send_from_directory(BASE_DIR,"register.html")

@app.route("/superadmin")
def superadmin_page(): return send_from_directory(BASE_DIR,"superadmin.html")

@app.route("/storage/uploads/logos/<filename>")
def serve_logo_static(filename):
    return send_from_directory(os.path.join(UPLOAD_FOLDER,"logos"),filename)

@app.route("/api/config/logo", methods=["POST"])
@require_auth
@require_role("admin")
def api_upload_logo():
    school_id = g.school_id
    if "logo" not in request.files:
        return jsonify({"ok":False,"error":"No logo file uploaded"}), 400
    f = request.files["logo"]
    if not f or not f.filename:
        return jsonify({"ok":False,"error":"No logo file uploaded"}), 400
    ext = f.filename.rsplit(".",1)[-1].lower() if "." in f.filename else ""
    if ext not in ALLOWED_LOGO_EXT:
        return jsonify({"ok":False,"error":"Logo must be an image (png, jpg, jpeg, gif, webp, svg)"}), 400
    raw = f.read()
    if len(raw) > 2*1024*1024:
        return jsonify({"ok":False,"error":"Logo must be smaller than 2MB"}), 400
    logo_mime = _mime_for_ext(ext)
    logo_b64  = base64.b64encode(raw).decode("ascii")
    set_config_val(school_id, "logo_data", logo_b64)
    set_config_val(school_id, "logo_mime", logo_mime)
    logo_path = f"api/logo/{school_id}"
    set_config_val(school_id, "logo_path", logo_path)
    return jsonify({"ok":True,"logo_path":logo_path})

@app.route("/api/logo/<int:school_id>")
def serve_logo_db(school_id):
    data = get_config_val(school_id, "logo_data", "")
    mime = get_config_val(school_id, "logo_mime", "image/png")
    if not data:
        return ("Not found", 404)
    try:
        raw = base64.b64decode(data)
    except Exception:
        return ("Not found", 404)
    return send_file(io.BytesIO(raw), mimetype=mime)

@app.route("/<path:filename>")
def serve_static(filename):
    if filename in _STATIC_FILES: return send_from_directory(BASE_DIR,filename)
    return ("Not found",404)


# ── STUDENT IMPORT (Excel/CSV) ─────────────────────────────────

import csv, io
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def hash_password_fast(pw):
    """Fast hash for temporary/bulk import passwords.
    These are throwaway — users must change on first login anyway.
    260,000 iterations x 4000 students = death. 100 iterations = fine.
    Iteration count is stored in the hash itself, so verify_password()
    still checks it correctly regardless of how it was hashed."""
    return hash_password(pw, iterations=100)

def _parse_import_file(file_obj, filename):
    """Parse uploaded Excel or CSV. Returns list of raw row dicts."""
    ext = filename.rsplit(".",1)[-1].lower() if "." in filename else ""
    rows = []
    if ext in ("xlsx","xls"):
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl not installed. Add it to requirements.txt")
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        headers = None
        for row in ws.iter_rows(values_only=True):
            # Skip fully empty rows
            if all(v is None for v in row): continue
            if headers is None:
                headers = [str(h).strip().lower() if h else "" for h in row]
                continue
            values = [str(v).strip() if v is not None else "" for v in row]
            row_dict = dict(zip(headers, values))
            # Skip completely empty rows
            if all(v == "" for v in values): continue
            rows.append(row_dict)
        wb.close()
    elif ext == "csv":
        text = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower(): str(v).strip() for k,v in row.items()})
    else:
        raise ValueError("Unsupported file type. Use .xlsx or .csv")
    return rows

def _normalize_col(row, *candidates):
    """Try multiple possible column name spellings — strip, lowercase, ignore spaces/underscores."""
    def clean(s): return s.strip().lower().replace(" ","").replace("_","").replace("-","")
    cleaned_row = {clean(k): v for k,v in row.items()}
    for c in candidates:
        key = clean(c)
        if key in cleaned_row and cleaned_row[key]:
            return str(cleaned_row[key]).strip()
    return ""

def _col_by_position(row, index):
    """Fallback: grab column by position index regardless of header name."""
    vals = list(row.values())
    if index < len(vals) and vals[index]:
        return str(vals[index]).strip()
    return ""

def _extract_fields(row):
    """
    Try named columns first. If name or class is still missing,
    fall back to positional: col0=name, col1=class, col2=stream, col3=phone.
    No conditions on format, length, or capitalization.
    """
    name         = _normalize_col(row,"name","student name","full name","jina","student","students","jina la mwanafunzi","mwanafunzi")
    class_name   = _normalize_col(row,"class_name","class","darasa","form","grade","class name","level","form name")
    stream_name  = _normalize_col(row,"stream_name","stream","mkondo","section","division","stream name","class stream")
    parent_phone = _normalize_col(row,"parent_phone","phone","parent phone","phone number","simu","contact","guardian phone","nambari","tel","telephone","mobile","simu ya mzazi","mzazi")
    # If name or class still missing — go positional, user probably has no headers or weird headers
    if not name:        name         = _col_by_position(row, 0)
    if not class_name:  class_name   = _col_by_position(row, 1)
    if not stream_name: stream_name  = _col_by_position(row, 2)
    if not parent_phone:parent_phone = _col_by_position(row, 3)
    return name, class_name, stream_name, parent_phone

def _build_class_map(school_id):
    """Return {class_name_lower: {stream_name_lower: (class_id, stream_id)}}"""
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT c.id,c.class_name,s.id,s.stream_name
                   FROM classes c LEFT JOIN streams s ON s.class_id=c.id AND s.school_id=c.school_id
                   WHERE c.school_id=%s""", (school_id,))
    rows = cur.fetchall(); cur.close(); con.close()
    cmap = {}
    for (cid, cname, sid, sname) in rows:
        ckey = cname.strip().lower()
        if ckey not in cmap: cmap[ckey] = {"_id": cid, "_streams": {}}
        if sname:
            skey = sname.strip().lower()
            cmap[ckey]["_streams"][skey] = (cid, sid)
    return cmap

_NUM_WORDS = {"one":"1","two":"2","three":"3","four":"4","five":"5","six":"6",
              "seven":"7","eight":"8","nine":"9","ten":"10","i":"1","ii":"2","iii":"3","iv":"4","v":"5"}

def _normalize_class_key(s):
    """Loose match key: lowercase, strip punctuation, collapse spaces, and turn
    spelled-out numbers ('form one') into digits ('form 1') so 'Form 1' and
    'FORM ONE' compare equal without the admin retyping anything."""
    if not s: return ""
    s = re.sub(r'[^a-z0-9\s]', ' ', str(s).strip().lower())
    s = re.sub(r'\s+', ' ', s).strip()
    return ' '.join(_NUM_WORDS.get(w, w) for w in s.split(' '))

def _resolve_class_stream(class_name, stream_name, cmap, mapping=None):
    """Resolve raw text from an import file to (class_id, stream_id, error).
    Tries an admin-supplied mapping first, then a loose normalized match,
    else returns an error string explaining what needs matching."""
    mapping = mapping or {"classes": {}, "streams": {}}
    resolved_class = (mapping["classes"].get(class_name.strip()) or class_name).strip()
    ckey = resolved_class.lower()
    if ckey not in cmap:
        norm = _normalize_class_key(resolved_class)
        ckey = next((k for k in cmap if _normalize_class_key(k) == norm), ckey)
    if ckey not in cmap:
        return None, None, f"Class '{class_name}' not found — match it in the mapping step"
    class_id = cmap[ckey]["_id"]; stream_id = None
    if stream_name:
        streams = cmap[ckey]["_streams"]
        resolved_stream = (mapping["streams"].get(f"{class_name.strip()}::{stream_name.strip()}") or stream_name).strip()
        skey = resolved_stream.lower()
        if skey not in streams:
            norm = _normalize_class_key(resolved_stream)
            skey = next((k for k in streams if _normalize_class_key(k) == norm), skey)
        if skey not in streams:
            return class_id, None, f"Stream '{stream_name}' not found in '{resolved_class}' — match it in the mapping step"
        class_id, stream_id = streams[skey]
    return class_id, stream_id, None

@app.route("/api/students/import/template", methods=["GET"])
@require_auth
@require_role("admin")
def api_import_template():
    """Download a pre-filled Excel template."""
    school_id = g.school_id
    if not OPENPYXL_AVAILABLE:
        # Fallback: return CSV template
        csv_content = "name,class_name,stream_name,parent_phone\nJuma Hassan,Form 1,A,0712345678\nFatuma Ally,Form 1,B,0754987654\n"
        return send_file(io.BytesIO(csv_content.encode()), as_attachment=True,
                        download_name="student_import_template.csv", mimetype="text/csv")
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Students"
    # Header row
    headers = ["name","class_name","stream_name","parent_phone"]
    header_labels = ["Student Full Name *","Class Name *","Stream Name (if any)","Parent Phone Number *"]
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1A6FA8")
    for col, (h, label) in enumerate(zip(headers, header_labels), 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    # Pull existing classes for reference
    cmap = _build_class_map(school_id)
    # Example rows
    examples = []
    for ckey, cval in list(cmap.items())[:3]:
        cname = ckey.title()
        if cval["_streams"]:
            for skey in list(cval["_streams"].keys())[:2]:
                examples.append([f"Example Student", cname, skey.title(), "07XXXXXXXXX"])
        else:
            examples.append(["Example Student", cname, "", "07XXXXXXXXX"])
    if not examples:
        examples = [
            ["Juma Hassan","Form 1","A","0712345678"],
            ["Fatuma Ally","Form 1","B","0754987654"],
            ["Emmanuel Peter","Form 2","","0622123456"],
        ]
    from openpyxl.styles import Font as F2
    for ri, row in enumerate(examples, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = F2(color="888888", italic=True)
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("STUDENT IMPORT INSTRUCTIONS",""),
        ("",""),
        ("Column","What to put"),
        ("name","Full name of the student (e.g. Juma Hassan Ally)"),
        ("class_name","Must match exactly a class in your school (e.g. Form 1, Form 2)"),
        ("stream_name","Stream if your class has streams (e.g. A, B, Science, Arts). Leave blank if no streams."),
        ("parent_phone","Parent or guardian phone number (e.g. 0712345678)"),
        ("",""),
        ("IMPORTANT",""),
        ("• Delete the example rows before importing",""),
        ("• class_name must exactly match your school classes",""),
        ("• stream_name must exactly match your stream names",""),
        ("• Do not change the column headers",""),
        ("• Save as .xlsx or .csv before uploading",""),
    ]
    from openpyxl.styles import Font as F3
    for ri, (a,b) in enumerate(instructions, 1):
        ws2.cell(row=ri, column=1, value=a)
        ws2.cell(row=ri, column=2, value=b)
        if ri == 1:
            ws2.cell(row=ri, column=1).font = F3(bold=True, size=13, color="1A6FA8")
        if ri == 3:
            ws2.cell(row=ri, column=1).font = F3(bold=True)
            ws2.cell(row=ri, column=2).font = F3(bold=True)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 50
    # Classes reference
    if cmap:
        ws3 = wb.create_sheet("Your Classes")
        ws3.cell(row=1,column=1,value="Class Name").font = F3(bold=True)
        ws3.cell(row=1,column=2,value="Streams").font = F3(bold=True)
        for ri,(ckey,cval) in enumerate(cmap.items(),2):
            ws3.cell(row=ri,column=1,value=ckey.title())
            streams = ", ".join(s.title() for s in cval["_streams"].keys())
            ws3.cell(row=ri,column=2,value=streams or "(no streams)")
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 30
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                    download_name="student_import_template.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/students/import/preview", methods=["POST"])
@require_auth
@require_role("admin")
def api_import_preview():
    school_id = g.school_id
    if "file" not in request.files:
        return jsonify({"ok":False,"error":"No file uploaded"}), 400
    f = request.files["file"]
    try:
        rows = _parse_import_file(f.stream, f.filename)
    except ValueError as e:
        return jsonify({"ok":False,"error":str(e)}), 400
    if not rows:
        return jsonify({"ok":False,"error":"File is empty or has no data rows"}), 400

    cmap = _build_class_map(school_id)
    unmatched_classes = {}
    unmatched_streams = {}
    preview = []
    for i, row in enumerate(rows):
        name, class_name, stream_name, parent_phone = _extract_fields(row)
        cid = sid = err = None
        if class_name:
            cid, sid, err = _resolve_class_stream(class_name, stream_name, cmap)
            if cid is None:
                unmatched_classes[class_name.strip().lower()] = class_name.strip()
            elif stream_name and sid is None and err:
                unmatched_streams[f"{class_name.strip()}::{stream_name.strip()}".lower()] = \
                    {"class_raw": class_name.strip(), "stream_raw": stream_name.strip()}
        if i < 10:
            issues = []
            if not name: issues.append("Missing name")
            if not class_name: issues.append("Missing class")
            elif cid is None: issues.append(f"Class '{class_name}' needs matching")
            elif stream_name and sid is None: issues.append(f"Stream '{stream_name}' needs matching")
            if not parent_phone: issues.append("Missing phone — will still import, no parent login created")
            preview.append({"row":i+2,"name":name,"class_name":class_name,"stream_name":stream_name,
                            "parent_phone":parent_phone,"issues":issues})

    return jsonify({"ok":True,"total_rows":len(rows),"preview":preview,
                    "columns_detected":list(rows[0].keys()) if rows else [],
                    "unmatched_classes": list(unmatched_classes.values()),
                    "unmatched_streams": list(unmatched_streams.values())})

def _write_credentials_xlsx(rows, path):
    """Write a workbook of parent username/one-time-password pairs for a completed import."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parent Login Credentials"

    headers = ["Row", "Student Name", "Class", "Stream", "Parent Phone", "Username", "Temporary Password"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1A6FA8", end_color="1A6FA8", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill

    for r in rows:
        ws.append([
            r["row"], r["name"], r["class_name"], r["stream_name"] or "-",
            r["parent_phone"], r["username"], r["password"],
        ])

    widths = [6, 26, 14, 12, 16, 22, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1,
            value="Note: this is a one-time password. Parents must change it after first login.").font = \
        Font(italic=True, color="888888")

    wb.save(path)


@app.route("/api/students/import/credentials/<path:filename>")
@require_auth
@require_role("admin")
def download_import_credentials(filename):
    """Serve a previously generated parent-credentials workbook for download."""
    filename = secure_filename(filename)
    path = os.path.join(IMPORT_EXPORT_FOLDER, filename)
    if not os.path.isfile(path):
        return jsonify({"ok": False, "error": "File not found or expired"}), 404
    return send_file(path, as_attachment=True,
                      download_name="parent_login_credentials.xlsx",
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/students/import", methods=["POST"])
@require_auth
@require_role("admin")
def api_import_students():
    """Full import: parse, validate, bulk insert."""
    school_id = g.school_id
    if "file" not in request.files:
        return jsonify({"ok":False,"error":"No file uploaded"}), 400
    f = request.files["file"]
    import traceback
    try:
        rows = _parse_import_file(f.stream, f.filename)
    except ValueError as e:
        return jsonify({"ok":False,"error":str(e)}),400
    except Exception as e:
        return jsonify({"ok":False,"error":"Parse failed: "+str(e)+" | "+traceback.format_exc()}), 500

    if not rows:
        return jsonify({"ok":False,"error":"File is empty"}), 400
    cmap     = _build_class_map(school_id)
    inserted = 0; skipped = []; errors = []; credentials = []; duplicates = 0
    con = get_db(); cur = con.cursor()

    # Snapshot of existing students for duplicate detection (name+class+stream+phone).
    # Also grows as we insert, so repeated rows within the same file are caught too —
    # and re-running an import on the same file becomes a safe no-op instead of creating dupes.
    cur.execute("SELECT LOWER(TRIM(name)), class_id, COALESCE(stream_id,0), phone_number "
                "FROM students WHERE school_id=%s", (school_id,))
    existing_students = set(cur.fetchall())
    cur.execute("SELECT COALESCE(MAX(school_student_no),0) FROM students WHERE school_id=%s", (school_id,))
    next_student_no = cur.fetchone()[0] + 1

    # Track which phone numbers already use each parent username, so we only
    # append "-{student_id}" when two DIFFERENT phone numbers sharing a
    # username would otherwise collide on the same last-4 password.
    cur.execute("""SELECT u.username, s.phone_number FROM users u JOIN students s ON u.student_id=s.id
                   WHERE u.school_id=%s AND u.role='parent'""", (school_id,))
    username_phone_map = {}
    for uname, ph in cur.fetchall():
        username_phone_map.setdefault(uname, []).append(ph or "")

    mapping_raw = request.form.get("mapping", "{}")
    try:
        mapping = json.loads(mapping_raw) if mapping_raw else {}
    except Exception:
        mapping = {}
    mapping.setdefault("classes", {}); mapping.setdefault("streams", {})
    flagged = []  # imported OK but missing a non-essential field — shown as a red dot in the UI

    for i, row in enumerate(rows):
        row_num = i + 2
        name, class_name, stream_name, parent_phone = _extract_fields(row)
        if not name:
            skipped.append({"row":row_num,"reason":"Missing name — can't import a student with no name","data":str(list(row.values())[:4])})
            continue
        if not class_name:
            skipped.append({"row":row_num,"reason":"Missing class","data":name})
            continue
        class_id, stream_id, resolve_err = _resolve_class_stream(class_name, stream_name, cmap, mapping)
        if resolve_err:
            skipped.append({"row":row_num,"reason":resolve_err,"data":name})
            continue
        phone_clean = (parent_phone or "").strip()
        dup_key = (name.strip().lower(), class_id, stream_id or 0, phone_clean)
        if dup_key in existing_students:
            skipped.append({"row":row_num,"reason":"Duplicate — same name, class, stream & phone already exist","data":name})
            duplicates += 1
            continue
        try:
            cur.execute("SAVEPOINT sp_student")
            cur.execute("""INSERT INTO students(school_id,name,class_id,stream_id,phone_number,school_student_no)
                           VALUES(%s,%s,%s,%s,%s,%s) RETURNING id""",
                        (school_id, name, class_id, stream_id, phone_clean or None, next_student_no))
            student_id = cur.fetchone()[0]
            next_student_no += 1
            if not phone_clean:
                reason = "Missing parent phone — no parent login was created"
                flagged.append({"row":row_num,"name":name,"reason":reason})
                cur.execute("UPDATE students SET flag_reason=%s WHERE id=%s",(reason,student_id))
                cur.execute("RELEASE SAVEPOINT sp_student")
                inserted += 1; existing_students.add(dup_key)
                if inserted % 200 == 0: con.commit()
                continue
            username_base = name.strip().lower().replace(" ","_")
            last4 = phone_clean[-4:]
            existing_phones_for_username = username_phone_map.get(username_base, [])
            needs_suffix = any(ph.strip() != phone_clean and ph.strip()[-4:] == last4 for ph in existing_phones_for_username)
            temp_pw = f"{last4}-{student_id}" if needs_suffix else last4
            cur.execute("INSERT INTO users(username,password,role,school_id,must_change_password,student_id) VALUES(%s,%s,'parent',%s,1,%s) ON CONFLICT(username,school_id) DO NOTHING",
                        (username_base,hash_password_fast(temp_pw),school_id,student_id))
            login_created = cur.rowcount > 0
            username_phone_map.setdefault(username_base, []).append(phone_clean)
            cur.execute("RELEASE SAVEPOINT sp_student")
            inserted += 1; existing_students.add(dup_key)
            if login_created:
                credentials.append({"row":row_num,"name":name,"class_name":class_name,"stream_name":stream_name,
                                    "parent_phone":phone_clean,"username":username_base,"password":temp_pw})
            else:
                reason = f"Login username '{username_base}' already taken by another student with the same name"
                flagged.append({"row":row_num,"name":name,"reason":reason})
                cur.execute("UPDATE students SET flag_reason=%s WHERE id=%s",(reason,student_id))
            if inserted % 200 == 0: con.commit()
        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT sp_student")
            errors.append({"row":row_num,"error":str(e),"data":name})
    con.commit(); cur.close(); con.close()

    credentials_file = None
    if credentials:
        fname = f"import_creds_{school_id}_{secrets.token_hex(8)}.xlsx"
        _write_credentials_xlsx(credentials, os.path.join(IMPORT_EXPORT_FOLDER, fname))
        credentials_file = f"/api/students/import/credentials/{fname}"

    return jsonify({"ok":True,"inserted":inserted,"skipped":len(skipped),"errors":len(errors),"duplicates":duplicates,
                    "skipped_details":skipped[:20],"error_details":errors[:20],
                    "flagged_details":flagged[:50],
                    "credentials_file":credentials_file,"credentials_count":len(credentials)})


with app.app_context():
    init_db()

if __name__=="__main__":
    app.run(debug=False,host="0.0.0.0",port=int(os.environ.get("PORT",5000)))

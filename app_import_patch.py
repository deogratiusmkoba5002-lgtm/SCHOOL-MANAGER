# HOW TO ADD IMPORT FEATURE TO YOUR app.py
#
# 1. Open your app.py on your local machine
# 2. Find this line near the bottom:
#       with app.app_context():
#           init_db()
# 3. Paste ALL the code below ABOVE that line
# 4. Add openpyxl to requirements.txt
# 5. Commit and push to Railway
#
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PASTE THIS ENTIRE BLOCK INTO YOUR app.py
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━


# ── STUDENT IMPORT (Excel/CSV) ─────────────────────────────────
import csv, io
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def hash_password_fast(pw):
    """Fast hash for temporary/bulk import passwords.
    These are throwaway — users must change on first login anyway.
    260,000 iterations x 4000 students = death. 100 iterations = fine."""
    salt = secrets.token_hex(16)
    dk   = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt.encode(), 100)
    return f"{salt}:{dk.hex()}"

def _parse_import_file(file_obj, filename):
    """Parse uploaded Excel or CSV. Returns list of raw row dicts."""
    ext = filename.rsplit(".",1)[-1].lower() if "." in filename else ""
    rows = []
    if ext in ("xlsx","xls"):
        if not OPENPYXL_AVAILABLE:
            raise ValueError("openpyxl not installed. Add it to requirements.txt")
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        ws = wb.active
        headers = None
        for row in ws.iter_rows(values_only=True):
            # Skip fully empty rows
            if all(v is None for v in row): continue
            if headers is None:
                headers = [str(h).strip().lower() if h else "" for h in row]
                continue
            values = [str(v).strip() if v is not None else "" for v in row]
            row_dict = dict(zip(headers, values))
            # Skip completely empty rows
            if all(v == "" for v in values): continue
            rows.append(row_dict)
        wb.close()
    elif ext == "csv":
        text = file_obj.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower(): str(v).strip() for k,v in row.items()})
    else:
        raise ValueError("Unsupported file type. Use .xlsx or .csv")
    return rows

def _normalize_col(row, *candidates):
    """Try multiple possible column name spellings — strip, lowercase, ignore spaces/underscores."""
    def clean(s): return s.strip().lower().replace(" ","").replace("_","").replace("-","")
    cleaned_row = {clean(k): v for k,v in row.items()}
    for c in candidates:
        key = clean(c)
        if key in cleaned_row and cleaned_row[key]:
            return str(cleaned_row[key]).strip()
    return ""

def _col_by_position(row, index):
    """Fallback: grab column by position index regardless of header name."""
    vals = list(row.values())
    if index < len(vals) and vals[index]:
        return str(vals[index]).strip()
    return ""

def _extract_fields(row):
    """
    Try named columns first. If name or class is still missing,
    fall back to positional: col0=name, col1=class, col2=stream, col3=phone.
    No conditions on format, length, or capitalization.
    """
    name         = _normalize_col(row,"name","student name","full name","jina","student","students","jina la mwanafunzi","mwanafunzi")
    class_name   = _normalize_col(row,"class_name","class","darasa","form","grade","class name","level","form name")
    stream_name  = _normalize_col(row,"stream_name","stream","mkondo","section","division","stream name","class stream")
    parent_phone = _normalize_col(row,"parent_phone","phone","parent phone","phone number","simu","contact","guardian phone","nambari","tel","telephone","mobile","simu ya mzazi","mzazi")
    # If name or class still missing — go positional, user probably has no headers or weird headers
    if not name:        name         = _col_by_position(row, 0)
    if not class_name:  class_name   = _col_by_position(row, 1)
    if not stream_name: stream_name  = _col_by_position(row, 2)
    if not parent_phone:parent_phone = _col_by_position(row, 3)
    return name, class_name, stream_name, parent_phone

def _build_class_map(school_id):
    """Return {class_name_lower: {stream_name_lower: (class_id, stream_id)}}"""
    con = get_db(); cur = con.cursor()
    cur.execute("""SELECT c.id,c.class_name,s.id,s.stream_name
                   FROM classes c LEFT JOIN streams s ON s.class_id=c.id AND s.school_id=c.school_id
                   WHERE c.school_id=%s""", (school_id,))
    rows = cur.fetchall(); cur.close(); con.close()
    cmap = {}
    for (cid, cname, sid, sname) in rows:
        ckey = cname.strip().lower()
        if ckey not in cmap: cmap[ckey] = {"_id": cid, "_streams": {}}
        if sname:
            skey = sname.strip().lower()
            cmap[ckey]["_streams"][skey] = (cid, sid)
    return cmap

@app.route("/api/students/import/template", methods=["GET"])
def api_import_template():
    """Download a pre-filled Excel template."""
    school_id = school_id_from_header()
    if not OPENPYXL_AVAILABLE:
        # Fallback: return CSV template
        csv_content = "name,class_name,stream_name,parent_phone\nJuma Hassan,Form 1,A,0712345678\nFatuma Ally,Form 1,B,0754987654\n"
        return send_file(io.BytesIO(csv_content.encode()), as_attachment=True,
                        download_name="student_import_template.csv", mimetype="text/csv")
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Students"
    # Header row
    headers = ["name","class_name","stream_name","parent_phone"]
    header_labels = ["Student Full Name *","Class Name *","Stream Name (if any)","Parent Phone Number *"]
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1A6FA8")
    for col, (h, label) in enumerate(zip(headers, header_labels), 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    # Pull existing classes for reference
    cmap = _build_class_map(school_id)
    # Example rows
    examples = []
    for ckey, cval in list(cmap.items())[:3]:
        cname = ckey.title()
        if cval["_streams"]:
            for skey in list(cval["_streams"].keys())[:2]:
                examples.append([f"Example Student", cname, skey.title(), "07XXXXXXXXX"])
        else:
            examples.append(["Example Student", cname, "", "07XXXXXXXXX"])
    if not examples:
        examples = [
            ["Juma Hassan","Form 1","A","0712345678"],
            ["Fatuma Ally","Form 1","B","0754987654"],
            ["Emmanuel Peter","Form 2","","0622123456"],
        ]
    from openpyxl.styles import Font as F2
    for ri, row in enumerate(examples, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = F2(color="888888", italic=True)
    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("STUDENT IMPORT INSTRUCTIONS",""),
        ("",""),
        ("Column","What to put"),
        ("name","Full name of the student (e.g. Juma Hassan Ally)"),
        ("class_name","Must match exactly a class in your school (e.g. Form 1, Form 2)"),
        ("stream_name","Stream if your class has streams (e.g. A, B, Science, Arts). Leave blank if no streams."),
        ("parent_phone","Parent or guardian phone number (e.g. 0712345678)"),
        ("",""),
        ("IMPORTANT",""),
        ("• Delete the example rows before importing",""),
        ("• class_name must exactly match your school classes",""),
        ("• stream_name must exactly match your stream names",""),
        ("• Do not change the column headers",""),
        ("• Save as .xlsx or .csv before uploading",""),
    ]
    from openpyxl.styles import Font as F3
    for ri, (a,b) in enumerate(instructions, 1):
        ws2.cell(row=ri, column=1, value=a)
        ws2.cell(row=ri, column=2, value=b)
        if ri == 1:
            ws2.cell(row=ri, column=1).font = F3(bold=True, size=13, color="1A6FA8")
        if ri == 3:
            ws2.cell(row=ri, column=1).font = F3(bold=True)
            ws2.cell(row=ri, column=2).font = F3(bold=True)
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 50
    # Classes reference
    if cmap:
        ws3 = wb.create_sheet("Your Classes")
        ws3.cell(row=1,column=1,value="Class Name").font = F3(bold=True)
        ws3.cell(row=1,column=2,value="Streams").font = F3(bold=True)
        for ri,(ckey,cval) in enumerate(cmap.items(),2):
            ws3.cell(row=ri,column=1,value=ckey.title())
            streams = ", ".join(s.title() for s in cval["_streams"].keys())
            ws3.cell(row=ri,column=2,value=streams or "(no streams)")
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 30
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                    download_name="student_import_template.xlsx",
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/students/import/preview", methods=["POST"])
def api_import_preview():
    """Parse file, return first 10 rows for preview without inserting."""
    school_id = school_id_from_header()
    if "file" not in request.files:
        return jsonify({"ok":False,"error":"No file uploaded"}), 400
    f = request.files["file"]
    try:
        rows = _parse_import_file(f.stream, f.filename)
    except ValueError as e:
        return jsonify({"ok":False,"error":str(e)}), 400
    if not rows:
        return jsonify({"ok":False,"error":"File is empty or has no data rows"}), 400
    cmap = _build_class_map(school_id)
    preview = []; warnings = []
    for i, row in enumerate(rows[:10]):
        name, class_name, stream_name, parent_phone = _extract_fields(row)
        errs = []
        if not name:        errs.append("Missing name")
        if not class_name:  errs.append("Missing class")
        if not parent_phone:errs.append("Missing phone")
        ckey = class_name.lower()
        if class_name and ckey not in cmap: errs.append(f"Class '{class_name}' not found")
        preview.append({"row":i+2,"name":name,"class_name":class_name,"stream_name":stream_name,
                        "parent_phone":parent_phone,"errors":errs})
    return jsonify({"ok":True,"total_rows":len(rows),"preview":preview,
                    "columns_detected":list(rows[0].keys()) if rows else []})

@app.route("/api/students/import", methods=["POST"])
def api_import_students():
    """Full import: parse, validate, bulk insert."""
    school_id = school_id_from_header()
    if "file" not in request.files:
        return jsonify({"ok":False,"error":"No file uploaded"}), 400
    f = request.files["file"]
    try:
        rows = _parse_import_file(f.stream, f.filename)
    except ValueError as e:
        return jsonify({"ok":False,"error":str(e)}), 400
    if not rows:
        return jsonify({"ok":False,"error":"File is empty"}), 400
    cmap     = _build_class_map(school_id)
    inserted = 0; skipped = []; errors = []
    con = get_db(); cur = con.cursor()
    for i, row in enumerate(rows):
        row_num      = i + 2
        name, class_name, stream_name, parent_phone = _extract_fields(row)
        if not name or not class_name or not parent_phone:
            skipped.append({"row":row_num,"reason":f"Missing: {'name' if not name else ''} {'class' if not class_name else ''} {'phone' if not parent_phone else ''}".strip(),"data":str(list(row.values())[:4])})
            continue
        ckey = class_name.strip().lower()
        if ckey not in cmap:
            skipped.append({"row":row_num,"reason":f"Class '{class_name}' not found — check spelling matches exactly","data":name})
            continue
        class_id  = cmap[ckey]["_id"]
        stream_id = None
        if stream_name:
            skey = stream_name.strip().lower()
            if skey in cmap[ckey]["_streams"]:
                class_id, stream_id = cmap[ckey]["_streams"][skey]
            else:
                skipped.append({"row":row_num,"reason":f"Stream '{stream_name}' not found in '{class_name}'","data":name})
                continue
        try:
            # Use savepoint so one failure doesn't abort the whole transaction
            cur.execute("SAVEPOINT sp_student")
            cur.execute("INSERT INTO students(school_id,name,class_id,stream_id,phone_number) VALUES(%s,%s,%s,%s,%s) RETURNING id",
                        (school_id, name, class_id, stream_id, parent_phone.strip()))
            student_id   = cur.fetchone()[0]
            username_base= name.strip().lower().replace(" ","_")
            last4        = parent_phone.strip()[-4:]
            cur.execute("SELECT id FROM students WHERE school_id=%s AND phone_number=%s AND id!=%s ORDER BY id",
                        (school_id, parent_phone.strip(), student_id))
            siblings = cur.fetchall()
            if siblings:
                cur.execute("SELECT id FROM students WHERE school_id=%s AND phone_number=%s ORDER BY id",(school_id,parent_phone.strip()))
                all_s=[r[0] for r in cur.fetchall()]
                try: idx=all_s.index(student_id)+1
                except ValueError: idx=len(all_s)+1
                temp_pw=f"{last4}-{idx}"
            else: temp_pw=last4
            cur.execute("SELECT username FROM users WHERE school_id=%s AND role='parent' AND username LIKE %s",
                        (school_id, username_base+"%"))
            existing=[r[0] for r in cur.fetchall()]
            final_user=username_base; counter=2
            while final_user in existing: final_user=f"{username_base}_{counter}"; counter+=1
            cur.execute("INSERT INTO users(username,password,role,school_id,must_change_password,student_id) VALUES(%s,%s,'parent',%s,1,%s) ON CONFLICT(username,school_id) DO NOTHING",
                        (final_user,hash_password_fast(temp_pw),school_id,student_id))
            cur.execute("RELEASE SAVEPOINT sp_student")
            inserted += 1
            # Commit every 200 students to avoid giant transactions
            if inserted % 200 == 0:
                con.commit()
        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT sp_student")
            errors.append({"row":row_num,"error":str(e),"data":name})
    con.commit(); cur.close(); con.close()
    return jsonify({"ok":True,"inserted":inserted,"skipped":len(skipped),"errors":len(errors),
                    "skipped_details":skipped[:20],"error_details":errors[:20]})
